"""Styled Excel workbook writer for SocietyHub month packs.

Blue header bands, per-owner and per-meter colour coding, shaded total rows and
red/green balance highlighting. Owner names and meter numbers are written verbatim.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = "1F3864"
BLUE = "2F5597"
BAND = "DCE6F5"
TOTAL = "C9D7EE"
TITLE = "0F172A"
RED = "C00000"
GREEN = "1F7A3D"

# Light, print-friendly tints cycled per owner / per meter. Generated so that a
# building with many owners never sees two of them share a colour.
def _tints(hues, light=0.94):
    import colorsys
    out = []
    for h in hues:
        r, g, b = colorsys.hls_to_rgb(h / 360, light, 0.85)
        out.append("%02X%02X%02X" % (int(r * 255), int(g * 255), int(b * 255)))
    return out


_HUES = [32, 130, 205, 285, 55, 185, 335, 260, 95, 15, 160, 240, 310, 75, 220, 355,
         115, 275, 45, 195, 300, 140, 20, 250, 85, 170, 320, 65, 230, 5]
OWNER_TINTS = _tints(_HUES)
METER_TINTS = _tints([(h + 18) % 360 for h in _HUES], light=0.90)
PAYER_TINTS = _tints([(h + 42) % 360 for h in _HUES], light=0.92)

MONEY = '#,##0.00'
NUM = '#,##0.00'
INT = '#,##0'

_thin = Side(style="thin", color="B9C4D4")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


class Palette:
    """Stable colour per key, in order of first appearance. `labels` maps key -> display name."""

    def __init__(self, tints):
        self.tints = tints
        self._map = {}
        self.labels = {}

    def fill(self, key, label=None):
        if key in (None, ""):
            return None
        if key not in self._map:
            self._map[key] = self.tints[len(self._map) % len(self.tints)]
        self.labels[key] = label or self.labels.get(key) or str(key)
        return PatternFill("solid", fgColor=self._map[key])

    def tint(self, key, label=None):
        """Hex tint for a key (registers it), or None."""
        self.fill(key, label)
        return self._map.get(key)


def new_book():
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def sheet(wb, name):
    ws = wb.create_sheet(name[:31])
    ws.sheet_view.showGridLines = False
    return ws


def title(ws, row, text, span=6, sub=None):
    ws.cell(row=row, column=1, value=text)
    c = ws.cell(row=row, column=1)
    c.font = Font(bold=True, size=14, color=TITLE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(span, 2))
    row += 1
    if sub:
        s = ws.cell(row=row, column=1, value=sub)
        s.font = Font(size=10, italic=True, color="5A6B85")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(span, 2))
        row += 1
    return row + 1


def section(ws, row, text, span=6):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=BLUE)
    c.alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(span, 2))
    ws.row_dimensions[row].height = 20
    return row + 1


def group_band(ws, row, spans, ncols):
    """A merged band above the header row, e.g. one 'Water Charges' cap over three columns."""
    for j in range(1, ncols + 1):
        ws.cell(row=row, column=j, value=None)
    for start, end, label in spans:
        c = ws.cell(row=row, column=start, value=label)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        for j in range(start + 1, end + 1):
            ws.cell(row=row, column=j).fill = PatternFill("solid", fgColor=BLUE)
            ws.cell(row=row, column=j).border = BORDER
        ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
    ws.row_dimensions[row].height = 18
    return row + 1


def table(ws, row, headers, rows, *, money_cols=(), int_cols=(), fills=None,
          signed_cols=(), total_row=None, widths=None):
    """Write a header band + data rows. `fills` is a list (one PatternFill or None per row)."""
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 30
    header_row = row
    row += 1

    for i, data in enumerate(rows):
        fill = (fills[i] if fills and i < len(fills) else None)
        for j, v in enumerate(data, start=1):
            c = ws.cell(row=row, column=j, value=v)
            c.border = BORDER
            c.font = Font(size=10)
            if fill:
                c.fill = fill
            if j in money_cols:
                c.number_format = MONEY
            elif j in int_cols:
                c.number_format = INT
            if j in signed_cols and isinstance(v, (int, float)):
                c.font = Font(size=10, bold=True, color=RED if v > 0 else (GREEN if v < 0 else "44506B"))
        row += 1

    if total_row is not None:
        for j, v in enumerate(total_row, start=1):
            c = ws.cell(row=row, column=j, value=v)
            c.fill = PatternFill("solid", fgColor=TOTAL)
            c.font = Font(bold=True, size=10)
            c.border = BORDER
            if j in money_cols:
                c.number_format = MONEY
            elif j in int_cols:
                c.number_format = INT
        row += 1

    if widths:
        for j, wd in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = wd
    else:
        for j, h in enumerate(headers, start=1):
            longest = max([len(str(h))] + [len(str(r[j - 1])) for r in rows if len(r) >= j])
            ws.column_dimensions[get_column_letter(j)].width = min(max(10, longest + 2), 34)

    return row + 1, header_row


def legend(ws, row, pairs, label="Legend"):
    row = section(ws, row, label, span=4)
    for k, v in pairs:
        kc = ws.cell(row=row, column=1, value=k)
        kc.font = Font(size=10, color="44506B")
        kc.fill = PatternFill("solid", fgColor=BAND)
        kc.border = BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        vc = ws.cell(row=row, column=4, value=v)
        vc.font = Font(size=10, bold=True)
        vc.border = BORDER
        if isinstance(v, float):
            vc.number_format = MONEY
        row += 1
    return row + 1


def colour_key(ws, row, palette, heading, only=None):
    """Show which tint belongs to which owner / meter / payer. `only` limits it to this sheet's keys."""
    keys = [k for k in palette._map if only is None or k in only]
    if not keys:
        return row
    row = section(ws, row, heading, span=4)
    for key in keys:
        c = ws.cell(row=row, column=1, value=palette.labels.get(key, str(key)))
        c.fill = PatternFill("solid", fgColor=palette._map[key])
        c.font = Font(size=10)
        c.border = BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
    return row + 1


def freeze(ws, cell):
    ws.freeze_panes = cell


def to_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
