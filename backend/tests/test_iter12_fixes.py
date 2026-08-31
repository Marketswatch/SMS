"""Iteration 12 — FIX ROUND verification for defects reported in iteration_11.json.

Covers: separate owner/payer/meter palettes, generated tints (>10 owners), meter tint
keyed on meter id, ?format validation (400), month validation (400), read-only
statement/export must NOT create period documents, Charges sheet freeze + colour key.

Run serially:  python -m pytest tests/test_iter12_fixes.py -n 0
"""
import io
import os

import pytest
import requests
from dotenv import dotenv_values
from openpyxl import load_workbook

fe = dotenv_values("/app/frontend/.env")
BASE = (os.environ.get("REACT_APP_BACKEND_URL") or fe.get("REACT_APP_BACKEND_URL")).rstrip("/")
API = f"{BASE}/api"
XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MIS_SHEETS = ["Water Reconciliation", "Water Usage (Meters)", "Tanker Purchases", "Charges", "Ledger"]


# ------------------------------------------------------------------ helpers
def rgb(cell):
    f = cell.fill
    if not f or f.fill_type != "solid":
        return None
    v = f.fgColor.rgb
    return v[-6:] if isinstance(v, str) else None


def find_row(ws, text, col=1, limit=400):
    for r in range(1, min(ws.max_row, limit) + 1):
        if str(ws.cell(row=r, column=col).value or "").strip() == text:
            return r
    return None


def key_block(ws, heading):
    """Return [(label, tint)] of a 'Colour key — x' block."""
    r = find_row(ws, heading)
    if r is None:
        return None
    out = []
    r += 1
    while r <= ws.max_row:
        v = ws.cell(row=r, column=1).value
        if v in (None, ""):
            break
        out.append((str(v), rgb(ws.cell(row=r, column=1))))
        r += 1
    return out


# ------------------------------------------------------------------ fixtures
@pytest.fixture(scope="session")
def client():
    s = requests.Session()
    r = s.post(f"{API}/auth/login", json={"email": "admin@societyhub.com", "password": "admin123"})
    assert r.status_code == 200, r.text
    tok = r.json().get("access_token") or r.json().get("token")
    if tok:
        s.headers.update({"Authorization": f"Bearer {tok}"})
    return s


@pytest.fixture(scope="session")
def prop(client):
    r = client.get(f"{API}/properties")
    assert r.status_code == 200, r.text
    props = r.json()
    named = [p for p in props if p["name"] == "Sunrise Residency"]
    assert named or props, "no maintenance property"
    return (named or props)[0]


MONTH = "2026-08"


@pytest.fixture(scope="session")
def wb(client, prop):
    r = client.get(f"{API}/mis/export",
                   params={"property_id": prop["id"], "month": MONTH, "format": "xlsx"})
    assert r.status_code == 200, r.text[:300]
    assert r.headers.get("content-type", "").startswith(XLSX_MEDIA)
    return load_workbook(io.BytesIO(r.content))


@pytest.fixture(scope="session")
def flats(client, prop):
    r = client.get(f"{API}/flats", params={"property_id": prop["id"]})
    assert r.status_code == 200
    return r.json()


# ================================================== 1. palettes separated
class TestPalettes:
    def test_sheets_unchanged(self, wb):
        assert wb.sheetnames == MIS_SHEETS

    def test_ledger_colour_key_lists_only_owners(self, wb, flats):
        block = key_block(wb["Ledger"], "Colour key — owner")
        assert block, "Ledger colour key block missing"
        labels = [b[0] for b in block]
        numbers = {f["number"] for f in flats}
        owners = {f["owner_name"] for f in flats}
        leaked = [l for l in labels if l in numbers or l.startswith("Flat ")]
        assert not leaked, f"flat numbers leaked into owner colour key: {leaked}"
        assert set(labels) <= owners, f"non-owner labels in owner key: {set(labels) - owners}"

    def test_recon_owner_tints_unique(self, wb):
        block = key_block(wb["Water Reconciliation"], "Colour key — owner")
        assert block
        tints = [b[1] for b in block]
        assert all(tints), "owner colour key rows have no fill"
        assert len(set(tints)) == len(tints), f"two owners share a tint: {block}"

    @pytest.mark.parametrize("sheetname", ["Tanker Purchases", "Charges"])
    def test_fronted_by_key_lists_flats(self, wb, sheetname, flats):
        block = key_block(wb[sheetname], "Colour key — fronted by")
        assert block, f"{sheetname} has no 'Colour key — fronted by' block"
        labels = [b[0] for b in block]
        assert all(l.startswith("Flat ") for l in labels), labels
        numbers = {f["number"] for f in flats}
        for l in labels:
            assert l.split("Flat ", 1)[1] in numbers, f"unknown flat in key: {l}"
        tints = [b[1] for b in block]
        assert len(set(tints)) == len(tints), f"payer tints collide: {block}"

    def test_owner_and_payer_palettes_are_independent(self, wb):
        owner_labels = {b[0] for b in key_block(wb["Ledger"], "Colour key — owner")}
        payer_labels = {b[0] for b in key_block(wb["Charges"], "Colour key — fronted by")}
        assert not (owner_labels & payer_labels)


# ============================================ 2. >10 owners / generated tints
@pytest.fixture(scope="module")
def scratch(client):
    """Scratch building with 12 flats/12 distinct owners + 2 same-label meters."""
    r = client.post(f"{API}/properties", json={"name": "ZZ QA Scratch Iter12", "address": "QA"})
    assert r.status_code == 200, r.text
    p = r.json()
    made = []
    for i in range(1, 13):
        fr = client.post(f"{API}/flats", json={
            "property_id": p["id"], "number": f"{900 + i}", "floor": "9",
            "owner_name": f"QA Owner {i:02d}"})
        assert fr.status_code == 200, fr.text
        made.append(fr.json())
    meters = []
    for f in made[:2]:
        mr = client.post(f"{API}/meters", json={"property_id": p["id"], "flat_id": f["id"],
                                                "label": "QA Dup Label 7", "opening": 0})
        assert mr.status_code == 200, mr.text
        meters.append(mr.json())
    yield {"prop": p, "flats": made, "meters": meters}
    assert client.delete(f"{API}/properties/{p['id']}").status_code == 200


@pytest.fixture(scope="module")
def scratch_wb(client, scratch):
    r = client.get(f"{API}/mis/export", params={"property_id": scratch["prop"]["id"],
                                                "month": MONTH, "format": "xlsx"})
    assert r.status_code == 200, r.text[:300]
    return load_workbook(io.BytesIO(r.content))


class TestManyOwners:
    def test_twelve_owners_all_distinct_tints(self, scratch_wb):
        block = key_block(scratch_wb["Water Reconciliation"], "Colour key — owner")
        assert block and len(block) == 12, block
        tints = [b[1] for b in block]
        assert all(tints)
        assert len(set(tints)) == 12, f"tints wrapped/collided: {sorted(tints)}"

    def test_row_fills_match_owner_key(self, scratch_wb):
        ws = scratch_wb["Water Reconciliation"]
        hdr = find_row(ws, "S.No")
        assert hdr
        seen = {}
        for r in range(hdr + 1, hdr + 13):
            owner = ws.cell(row=r, column=4).value
            seen[owner] = rgb(ws.cell(row=r, column=4))
        assert len(seen) == 12
        assert len(set(seen.values())) == 12, f"duplicate row tints: {seen}"

    # ============ 3. meter tint keyed on meter id, label displayed verbatim
    def test_same_label_meters_get_different_tints(self, scratch_wb):
        ws = scratch_wb["Water Usage (Meters)"]
        hdr = find_row(ws, "S.No")
        assert hdr
        rows = []
        for r in range(hdr + 1, hdr + 3):
            rows.append((ws.cell(row=r, column=5).value, rgb(ws.cell(row=r, column=5))))
        assert len(rows) == 2, rows
        assert rows[0][0] == rows[1][0] == "QA Dup Label 7", rows
        assert rows[0][1] and rows[1][1]
        assert rows[0][1] != rows[1][1], f"same-label meters share a tint: {rows}"

    def test_meter_colour_key_shows_label_verbatim(self, scratch_wb):
        block = key_block(scratch_wb["Water Usage (Meters)"], "Colour key — meter")
        assert block and len(block) == 2, block
        assert [b[0] for b in block] == ["QA Dup Label 7", "QA Dup Label 7"]
        assert block[0][1] != block[1][1]


# ================================================== 4. format validation
class TestFormatValidation:
    @pytest.mark.parametrize("bad", ["bogus", "xls", "PDF", "", "html"])
    def test_mis_export_bad_format(self, client, prop, bad):
        r = client.get(f"{API}/mis/export",
                       params={"property_id": prop["id"], "month": MONTH, "format": bad})
        assert r.status_code == 400, f"{bad} -> {r.status_code} {r.headers.get('content-type')}"

    @pytest.mark.parametrize("fmt,ct", [("csv", "text/csv"), ("pdf", "application/pdf"),
                                        ("xlsx", XLSX_MEDIA)])
    def test_mis_export_good_format(self, client, prop, fmt, ct):
        r = client.get(f"{API}/mis/export",
                       params={"property_id": prop["id"], "month": MONTH, "format": fmt})
        assert r.status_code == 200, r.text[:200]
        assert ct in r.headers.get("content-type", "")

    @pytest.mark.parametrize("bad", ["bogus", "xlsx", "xls"])
    def test_annual_export_bad_format(self, client, prop, bad):
        r = client.get(f"{API}/annual/export",
                       params={"property_id": prop["id"], "year": 2026, "format": bad})
        assert r.status_code == 400, f"{bad} -> {r.status_code}"

    @pytest.mark.parametrize("fmt,ct", [("csv", "text/csv"), ("pdf", "application/pdf")])
    def test_annual_export_good_format(self, client, prop, fmt, ct):
        r = client.get(f"{API}/annual/export",
                       params={"property_id": prop["id"], "year": 2026, "format": fmt})
        assert r.status_code == 200, r.text[:200]
        assert ct in r.headers.get("content-type", "")

    @pytest.mark.parametrize("bad", ["bogus", "xls", "doc"])
    def test_rentals_export_bad_format(self, client, bad):
        r = client.get(f"{API}/rentals/export", params={"month": MONTH, "format": bad})
        assert r.status_code == 400, f"{bad} -> {r.status_code}"

    @pytest.mark.parametrize("fmt,ct", [("csv", "text/csv"), ("pdf", "application/pdf"),
                                        ("xlsx", XLSX_MEDIA)])
    def test_rentals_export_good_format(self, client, fmt, ct):
        r = client.get(f"{API}/rentals/export", params={"month": MONTH, "format": fmt})
        assert r.status_code == 200, r.text[:200]
        assert ct in r.headers.get("content-type", "")


# ================================================== 5. month validation
class TestMonthValidation:
    @pytest.mark.parametrize("bad", ["bad", "2026-13", "2026-00", "202608", "2026-8", ""])
    def test_mis_export_bad_month(self, client, prop, bad):
        r = client.get(f"{API}/mis/export",
                       params={"property_id": prop["id"], "month": bad, "format": "xlsx"})
        assert r.status_code == 400, f"{bad} -> {r.status_code}"

    @pytest.mark.parametrize("bad", ["bad", "2026-13", "2026-00", "2026-8"])
    def test_statement_bad_month(self, client, prop, bad):
        r = client.get(f"{API}/statement", params={"property_id": prop["id"], "month": bad})
        assert r.status_code == 400, f"{bad} -> {r.status_code}"

    def test_reads_do_not_create_periods(self, client, prop):
        pid = prop["id"]
        before = {p["month"] for p in client.get(f"{API}/periods", params={"property_id": pid}).json()}
        empty = "2029-11"
        assert empty not in before
        for _ in range(3):
            assert client.get(f"{API}/statement", params={"property_id": pid, "month": empty}).status_code == 200
            assert client.get(f"{API}/mis/export", params={"property_id": pid, "month": empty,
                                                           "format": "xlsx"}).status_code == 200
            assert client.get(f"{API}/overview", params={"property_id": pid, "month": empty}).status_code in (200, 404)
        after = {p["month"] for p in client.get(f"{API}/periods", params={"property_id": pid}).json()}
        assert after == before, f"read-only calls created periods: {after - before}"

    def test_writes_still_create_and_lock_periods(self, client, scratch):
        pid = scratch["prop"]["id"]
        m = "2029-12"
        before = {p["month"] for p in client.get(f"{API}/periods", params={"property_id": pid}).json()}
        assert m not in before
        r = client.post(f"{API}/charges", json={"property_id": pid, "month": m,
                                                "charge_type": "misc",
                                                "description": "TEST_iter12", "amount": 120,
                                                "payer_type": "owner"})
        assert r.status_code == 200, r.text
        cid = r.json()["id"]
        after = {p["month"] for p in client.get(f"{API}/periods", params={"property_id": pid}).json()}
        assert m in after, "writing a charge did not create the period"
        assert client.delete(f"{API}/charges/{cid}").status_code == 200


# ================================================== 6. Charges sheet polish
class TestChargesSheet:
    def test_charges_frozen_panes(self, wb):
        assert wb["Charges"].freeze_panes, "Charges sheet still has no frozen panes"

    def test_all_sheets_frozen(self, wb):
        missing = [n for n in MIS_SHEETS if not wb[n].freeze_panes]
        assert not missing, missing

    def test_charges_sections_and_totals(self, wb):
        ws = wb["Charges"]
        texts = [str(ws.cell(row=r, column=1).value or "") for r in range(1, ws.max_row + 1)]
        totals = [c.coordinate for row in ws.iter_rows() for c in row
                  if str(c.value).strip() == "TOTAL"]
        assert any(t.startswith("Recurring —") for t in texts), texts[:20]
        assert any(t.startswith("One-time") for t in texts), texts[:20]
        assert len(totals) >= 2, totals

    def test_navy_headers_everywhere(self, wb):
        for n in MIS_SHEETS:
            ws = wb[n]
            hdr = find_row(ws, "S.No")
            assert hdr, f"{n} has no S.No header"
            c = ws.cell(row=hdr, column=1)
            assert rgb(c) == "1F3864", f"{n} header fill {rgb(c)}"
            assert c.font.bold and c.font.color.rgb[-6:] == "FFFFFF"


# ============================== 7. month close & carry-forward (build_statement ensure flag)
@pytest.fixture(scope="module")
def closable(client):
    r = client.post(f"{API}/properties", json={"name": "ZZ QA Close Iter12", "address": "QA"})
    assert r.status_code == 200, r.text
    p = r.json()
    f1 = client.post(f"{API}/flats", json={"property_id": p["id"], "number": "701",
                                           "floor": "7", "owner_name": "QA Close Owner A"}).json()
    f2 = client.post(f"{API}/flats", json={"property_id": p["id"], "number": "702",
                                           "floor": "7", "owner_name": "QA Close Owner B"}).json()
    m1 = client.post(f"{API}/meters", json={"property_id": p["id"], "flat_id": f1["id"],
                                            "label": "QA M701", "opening": 100}).json()
    m2 = client.post(f"{API}/meters", json={"property_id": p["id"], "flat_id": f2["id"],
                                            "label": "QA M702", "opening": 200}).json()
    m = "2027-05"
    assert client.put(f"{API}/readings", json={"property_id": p["id"], "month": m, "readings": [
        {"meter_id": m1["id"], "opening": 100, "closing": 1100},
        {"meter_id": m2["id"], "opening": 200, "closing": 700}]}).status_code == 200
    assert client.post(f"{API}/tankers", json={"property_id": p["id"], "month": m,
                                               "date": f"{m}-03", "qty_sump": 6000, "amount": 2400,
                                               "payer_flat_id": f1["id"], "payer_type": "owner",
                                               "tips_amount": 100,
                                               "tips_payer_flat_id": f1["id"]}).status_code == 200
    assert client.post(f"{API}/charges", json={"property_id": p["id"], "month": m,
                                               "charge_type": "security", "description": "TEST_iter12 guard",
                                               "amount": 1000, "payer_type": "owner"}).status_code == 200
    assert client.post(f"{API}/payments", json={"property_id": p["id"], "month": m, "flat_id": f2["id"],
                                                "amount": 500, "date": f"{m}-10"}).status_code == 200
    yield {"prop": p, "flats": [f1, f2], "meters": [m1, m2], "month": m}
    assert client.delete(f"{API}/properties/{p['id']}").status_code == 200


class TestMonthClose:
    def test_payment_direction_is_constrained(self, client, closable):
        r = client.post(f"{API}/payments", json={"property_id": closable["prop"]["id"],
                                                 "month": closable["month"],
                                                 "flat_id": closable["flats"][0]["id"],
                                                 "amount": 10, "date": f"{closable['month']}-11",
                                                 "direction": "bogus"})
        assert r.status_code == 422, r.status_code

    def test_close_month_carries_forward(self, client, closable):
        pid, m = closable["prop"]["id"], closable["month"]
        before = client.get(f"{API}/statement", params={"property_id": pid, "month": m})
        assert before.status_code == 200, before.text
        nets = {r["flat_id"]: r["net"] for r in before.json()["rows"]}
        assert any(abs(v) > 0 for v in nets.values()), nets

        r = client.post(f"{API}/periods/reset", params={"property_id": pid, "month": m})
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["locked_month"] == m and body["new_month"] == "2027-06"
        assert {k: round(v, 2) for k, v in body["carry_in"].items()} == {k: round(v, 2) for k, v in nets.items()}

        # locked month rejects writes
        w = client.post(f"{API}/charges", json={"property_id": pid, "month": m, "charge_type": "misc",
                                                "description": "TEST_iter12 locked", "amount": 5})
        assert w.status_code == 423, w.status_code

        # next month statement shows the brought-forward amounts + carried meter openings
        nxt = client.get(f"{API}/statement", params={"property_id": pid, "month": "2027-06"})
        assert nxt.status_code == 200, nxt.text
        data = nxt.json()
        assert data["status"] == "open"
        for row in data["rows"]:
            assert round(row["carry_in"], 2) == round(nets[row["flat_id"]], 2), row
        openings = {mm["label"]: mm["opening"] for mm in data["meters"]}
        assert openings == {"QA M701": 1100, "QA M702": 700}, openings

        # locked month statement/export still readable
        assert client.get(f"{API}/statement", params={"property_id": pid, "month": m}).status_code == 200
        x = client.get(f"{API}/mis/export", params={"property_id": pid, "month": m, "format": "xlsx"})
        assert x.status_code == 200 and x.headers["content-type"].startswith(XLSX_MEDIA)

    def test_double_reset_rejected(self, client, closable):
        r = client.post(f"{API}/periods/reset",
                        params={"property_id": closable["prop"]["id"], "month": closable["month"]})
        assert r.status_code == 423, r.status_code
