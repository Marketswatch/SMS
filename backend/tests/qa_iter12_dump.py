import io, os
import requests
from dotenv import dotenv_values
from openpyxl import load_workbook

fe = dotenv_values("/app/frontend/.env")
API = (os.environ.get("REACT_APP_BACKEND_URL") or fe.get("REACT_APP_BACKEND_URL")).rstrip("/") + "/api"
s = requests.Session()
r = s.post(f"{API}/auth/login", json={"email": "admin@societyhub.com", "password": "admin123"})
tok = r.json().get("access_token") or r.json().get("token")
if tok:
    s.headers.update({"Authorization": f"Bearer {tok}"})
prop = [p for p in s.get(f"{API}/properties").json() if p["name"] == "Sunrise Residency"][0]
wb = load_workbook(io.BytesIO(s.get(f"{API}/mis/export", params={"property_id": prop["id"], "month": "2026-08", "format": "xlsx"}).content))
for name in ("Water Usage (Meters)", "Charges"):
    ws = wb[name]
    print("=====", name, "freeze:", ws.freeze_panes)
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40)):
        vals = [(c.value, c.fill.fgColor.rgb[-6:] if c.fill and c.fill.fill_type == "solid" else "") for c in row if c.value not in (None, "")]
        if vals:
            print(row[0].row, vals)
