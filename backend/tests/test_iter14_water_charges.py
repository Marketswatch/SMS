"""Iteration 14 — water reconciliation owner-report changes.

(a) 'Water Charges' merged group band over Metered | Non-Metered (in storage) | Total Water cost
(b) four payment statuses: Prepaid / Paid / Pending / Excess Paid Back
(c) colour coding per FLAT (both meters of a flat share one tint)
(d) meters report last column renamed 'Total Amount'
(e) readings write by meter_id (closing reading bug)
(f) one page per report, no colour-key blocks

Run serially:  python -m pytest tests/test_iter14_water_charges.py -n 0
"""
import io
import os
import re

import fitz
import pytest
import requests
from dotenv import dotenv_values
from openpyxl import load_workbook

fe = dotenv_values("/app/frontend/.env")
BASE = (os.environ.get("REACT_APP_BACKEND_URL") or fe.get("REACT_APP_BACKEND_URL")).rstrip("/")
API = f"{BASE}/api"
MONTH = "2026-08"
STATUS_LABELS = {"Prepaid", "Paid", "Pending", "Excess Paid Back"}


@pytest.fixture(scope="module")
def client():
    s = requests.Session()
    r = s.post(f"{API}/auth/login", json={"email": "admin@societyhub.com", "password": "admin123"})
    assert r.status_code == 200, r.text
    tok = r.json().get("access_token") or r.json().get("token")
    if tok:
        s.headers.update({"Authorization": f"Bearer {tok}"})
    return s


@pytest.fixture(scope="module")
def pid(client):
    props = client.get(f"{API}/properties").json()
    named = [p for p in props if p["name"] == "Sunrise Residency"]
    return (named or props)[0]["id"]


def stmt_of(client, pid):
    r = client.get(f"{API}/statement", params={"property_id": pid, "month": MONTH})
    assert r.status_code == 200, r.text
    return r.json()


# ------------------------------------------------------------------ (e) readings
class TestReadingsByMeterId:
    def test_each_closing_value_lands_on_its_own_meter(self, client, pid):
        orig = client.get(f"{API}/readings", params={"property_id": pid, "month": MONTH}).json()
        assert len(orig) >= 2, f"need >=2 meters, got {len(orig)}"
        try:
            # distinct closing per meter, deliberately in REVERSE order (sorted-index bug would swap)
            wanted = {}
            payload = []
            for i, r in enumerate(reversed(orig)):
                closing = float(r["opening"]) + 100 + i
                wanted[r["meter_id"]] = closing
                payload.append({"meter_id": r["meter_id"], "opening": float(r["opening"]),
                                "closing": closing, "media": r.get("media") or []})
            res = client.put(f"{API}/readings", json={"property_id": pid, "month": MONTH,
                                                     "readings": payload})
            assert res.status_code == 200, res.text
            for row in res.json():
                assert row["closing"] == pytest.approx(wanted[row["meter_id"]]), \
                    f"PUT response mismatched for meter {row['label']}"
            got = client.get(f"{API}/readings", params={"property_id": pid, "month": MONTH}).json()
            for row in got:
                assert row["closing"] == pytest.approx(wanted[row["meter_id"]]), \
                    f"persisted closing wrong for meter {row['label']}"
            # consumption must follow on the right meter
            meters = {m["meter_id"] if "meter_id" in m else m.get("id"): m
                      for m in stmt_of(client, pid)["meters"]}
            for mid, closing in wanted.items():
                m = meters.get(mid)
                if m:
                    assert m["consumption"] == pytest.approx(closing - float(m["opening"]), abs=0.02)
        finally:
            client.put(f"{API}/readings", json={
                "property_id": pid, "month": MONTH,
                "readings": [{"meter_id": r["meter_id"], "opening": float(r["opening"]),
                              "closing": r["closing"], "media": r.get("media") or []} for r in orig]})

    def test_media_saved_against_correct_meter(self, client, pid):
        orig = client.get(f"{API}/readings", params={"property_id": pid, "month": MONTH}).json()
        target = orig[-1]
        try:
            payload = [{"meter_id": r["meter_id"], "opening": float(r["opening"]), "closing": r["closing"],
                        "media": ([{"file_id": "TEST_media", "name": "TEST.png"}]
                                  if r["meter_id"] == target["meter_id"] else (r.get("media") or []))}
                       for r in orig]
            res = client.put(f"{API}/readings", json={"property_id": pid, "month": MONTH, "readings": payload})
            assert res.status_code == 200, res.text
            for row in res.json():
                names = [m.get("name") for m in row.get("media") or []]
                if row["meter_id"] == target["meter_id"]:
                    assert "TEST.png" in names
                else:
                    assert "TEST.png" not in names
        finally:
            client.put(f"{API}/readings", json={
                "property_id": pid, "month": MONTH,
                "readings": [{"meter_id": r["meter_id"], "opening": float(r["opening"]),
                              "closing": r["closing"], "media": r.get("media") or []} for r in orig]})


# ------------------------------------------------------------------ (b) statuses
class TestPaymentStatus:
    def test_only_four_statuses_and_logic(self, client, pid):
        rows = stmt_of(client, pid)["rows"]
        assert rows
        for r in rows:
            ps = r["payment_status"]
            assert ps in {"prepaid", "paid", "pending", "excess_paid_back"}, ps
            net, payout = r["net"], r.get("payouts", 0)
            if net < -0.005:
                assert ps == "prepaid"
            elif net > 0.005:
                assert ps == "pending"
            else:
                assert ps == ("excess_paid_back" if payout > 0.005 else "paid")

    def test_payout_flips_prepaid_to_excess_paid_back(self, client, pid):
        rows = stmt_of(client, pid)["rows"]
        credit = [r for r in rows if r["net"] < -0.005]
        assert credit, "expected at least one flat in credit (prepaid)"
        target = credit[0]
        amount = round(-target["net"], 2)
        pay = client.post(f"{API}/payments", json={
            "property_id": pid, "month": MONTH, "flat_id": target["flat_id"], "amount": amount,
            "date": f"{MONTH}-28", "direction": "payout", "payer_type": "owner",
            "notes": "TEST_iter14 refund of excess"})
        assert pay.status_code in (200, 201), pay.text
        pay_id = pay.json().get("id")
        try:
            after = next(r for r in stmt_of(client, pid)["rows"] if r["flat_id"] == target["flat_id"])
            assert after["net"] == pytest.approx(0, abs=0.02), after["net"]
            assert after["payment_status"] == "excess_paid_back", after["payment_status"]
            # Excel
            xb = client.get(f"{API}/mis/export", params={"property_id": pid, "month": MONTH,
                                                        "format": "xlsx"}).content
            wb = load_workbook(io.BytesIO(xb))
            ws = wb["Water Reconciliation"]
            labels = [c.value for r in ws.iter_rows() for c in r
                      if isinstance(c.value, str) and c.value in STATUS_LABELS]
            assert "Excess Paid Back" in labels, labels
            # PDF
            pdf = client.get(f"{API}/reports/pack", params={"property_id": pid, "month": MONTH,
                                                           "report": "reconciliation",
                                                           "format": "pdf"}).content
            text = "".join(p.get_text() for p in fitz.open(stream=pdf, filetype="pdf"))
            assert "Excess Paid Back" in text
        finally:
            if pay_id:
                assert client.delete(f"{API}/payments/{pay_id}").status_code in (200, 204)
        back = next(r for r in stmt_of(client, pid)["rows"] if r["flat_id"] == target["flat_id"])
        assert back["payment_status"] == "prepaid"

    def test_no_legacy_status_labels_in_exports(self, client, pid):
        pdf = client.get(f"{API}/reports/pack", params={"property_id": pid, "month": MONTH,
                                                        "format": "pdf"}).content
        text = "".join(p.get_text() for p in fitz.open(stream=pdf, filetype="pdf"))
        assert "Partial" not in text
        assert not re.search(r"\bSettled\b", text), "legacy 'Settled' label in PDF"
        xb = client.get(f"{API}/mis/export", params={"property_id": pid, "month": MONTH,
                                                     "format": "xlsx"}).content
        wb = load_workbook(io.BytesIO(xb))
        for ws in wb.worksheets:
            vals = [c.value for r in ws.iter_rows() for c in r if isinstance(c.value, str)]
            assert "Partial" not in vals, ws.title
            assert "Settled" not in vals, ws.title
        csv_text = client.get(f"{API}/mis/export", params={"property_id": pid, "month": MONTH,
                                                           "format": "csv"}).text
        assert "Partial" not in csv_text
        assert ",Settled" not in csv_text


# ------------------------------------------------------------ (a)+(d) headers
class TestGroupBandAndHeaders:
    def test_xlsx_group_band_and_headers(self, client, pid):
        xb = client.get(f"{API}/mis/export", params={"property_id": pid, "month": MONTH,
                                                     "format": "xlsx"}).content
        ws = load_workbook(io.BytesIO(xb))["Water Reconciliation"]
        band = [c for r in ws.iter_rows() for c in r if c.value == "Water Charges"]
        assert len(band) == 1, "expected exactly one 'Water Charges' band cell"
        cell = band[0]
        merges = [str(m) for m in ws.merged_cells.ranges]
        target = next((m for m in ws.merged_cells.ranges
                       if m.min_row == cell.row and m.min_col == cell.column), None)
        assert target is not None, f"band cell {cell.coordinate} is not merged; merges={merges}"
        assert (target.min_col, target.max_col) == (5, 7), f"band spans {target}"
        hdr = [c.value for c in ws[cell.row + 1]]
        assert hdr[4:7] == ["Metered", "Non-Metered (in storage)", "Total Water cost"], hdr[4:7]

    def test_csv_group_band(self, client, pid):
        text = client.get(f"{API}/mis/export", params={"property_id": pid, "month": MONTH,
                                                       "format": "csv"}).text
        lines = text.splitlines()
        band_idx = next((i for i, ln in enumerate(lines) if ln.count("Water Charges") == 3), None)
        assert band_idx is not None, "no CSV band row with three 'Water Charges' cells"
        cells = lines[band_idx].split(",")
        assert cells[4:7] == ["Water Charges"] * 3, cells[:8]
        hdr = lines[band_idx + 1].split(",")
        assert hdr[4] == "Metered" and hdr[6] == "Total Water cost", hdr[:8]

    def test_pdf_group_band(self, client, pid):
        pdf = client.get(f"{API}/reports/pack", params={"property_id": pid, "month": MONTH,
                                                        "report": "reconciliation",
                                                        "format": "pdf"}).content
        doc = fitz.open(stream=pdf, filetype="pdf")
        page = doc[0]
        band = page.search_for("Water Charges")
        assert band, "no 'Water Charges' cap in reconciliation PDF"
        metered = page.search_for("Metered")
        assert metered, "no 'Metered' header"
        # the cap must sit above the column headers
        assert band[0].y1 <= min(m.y0 for m in metered) + 2

    def test_total_amount_heading(self, client, pid):
        xb = client.get(f"{API}/mis/export", params={"property_id": pid, "month": MONTH,
                                                     "format": "xlsx"}).content
        ws = load_workbook(io.BytesIO(xb))["Water Usage (Meters)"]
        headers = []
        for r in ws.iter_rows():
            vals = [c.value for c in r]
            if "Consumed units" in vals:
                headers = vals
                break
        assert headers, "meters header row not found"
        assert "Total Amount" in headers, f"meters sheet last column headers: {headers}"
        assert "Combined (per flat)" not in headers, "old heading still present in Excel"

        csv_text = client.get(f"{API}/mis/export", params={"property_id": pid, "month": MONTH,
                                                           "format": "csv"}).text
        assert "Total Amount" in csv_text
        assert "Combined (per flat)" not in csv_text

        pdf = client.get(f"{API}/reports/pack", params={"property_id": pid, "month": MONTH,
                                                        "report": "meters", "format": "pdf"}).content
        text = "".join(p.get_text() for p in fitz.open(stream=pdf, filetype="pdf"))
        assert "Combined" not in text, "PDF meters report still says 'Combined'"
        assert "Total" in text and "Amount" in text


# ------------------------------------------------------------------ (c) colours
class TestPerFlatColours:
    @pytest.fixture(scope="class")
    def extra_meter(self, client, pid):
        rows = stmt_of(client, pid)["rows"]
        flat = rows[0]
        r = client.post(f"{API}/meters", json={"property_id": pid, "flat_id": flat["flat_id"],
                                               "label": "TEST_M2", "opening": 10})
        assert r.status_code in (200, 201), r.text
        mid = r.json()["id"]
        yield flat, mid
        assert client.delete(f"{API}/meters/{mid}").status_code in (200, 204)

    def test_xlsx_one_tint_per_flat_across_sheets(self, client, pid, extra_meter):
        flat, _mid = extra_meter
        xb = client.get(f"{API}/mis/export", params={"property_id": pid, "month": MONTH,
                                                     "format": "xlsx"}).content
        wb = load_workbook(io.BytesIO(xb))

        def flat_colours(sheet, flat_col):
            out = {}
            for row in sheet.iter_rows():
                v = row[flat_col].value
                if v in (None, "", "TOTAL"):
                    continue
                key = str(v)
                fill = row[0].fill
                rgb = fill.start_color.rgb if fill and fill.start_color else None
                if isinstance(rgb, str) and rgb not in ("00000000",):
                    out.setdefault(key, []).append(rgb)
            return out

        recon = flat_colours(wb["Water Reconciliation"], 1)
        meters = flat_colours(wb["Water Usage (Meters)"], 1)
        # remove header/title artifacts by keeping keys that look like flat numbers
        recon = {k: v for k, v in recon.items() if re.fullmatch(r"[A-Za-z0-9\-/ ]{1,10}", k)
                 and k not in ("Flat No.", "House")}
        meters = {k: v for k, v in meters.items() if k not in ("Flat No.", "House")}

        # per flat one colour
        for k, v in recon.items():
            assert len(set(v)) == 1, f"recon flat {k} has multiple tints {set(v)}"
        for k, v in meters.items():
            assert len(set(v)) == 1, f"meters flat {k} has multiple tints {set(v)}"
        # two meters of the new flat share the tint
        target = str(flat["flat_number"])
        assert target in meters and len(meters[target]) >= 2, \
            f"flat {target} should have 2 meter rows, got {meters.get(target)}"
        # different flats differ
        uniq = {k: set(v).pop() for k, v in meters.items()}
        assert len(set(uniq.values())) == len(uniq), f"tints collide across flats: {uniq}"
        # same flat -> same colour between the two sheets
        shared = set(recon) & set(uniq)
        for k in shared:
            assert set(recon[k]).pop() == uniq[k], f"flat {k} tint differs between sheets"

    def test_ledger_sheet_uses_same_per_flat_tints(self, client, pid):
        xb = client.get(f"{API}/mis/export", params={"property_id": pid, "month": MONTH,
                                                     "format": "xlsx"}).content
        wb = load_workbook(io.BytesIO(xb))
        ledger, recon = wb["Ledger"], wb["Water Reconciliation"]

        def tints(sheet, flat_col, header):
            out, started = {}, False
            for row in sheet.iter_rows():
                vals = [c.value for c in row]
                if header in vals:
                    started = True
                    continue
                if not started:
                    continue
                v = row[flat_col].value
                if v in (None, "", "TOTAL"):
                    continue
                rgb = row[0].fill.start_color.rgb if row[0].fill else None
                if isinstance(rgb, str) and rgb != "00000000":
                    out.setdefault(str(v), set()).add(rgb)
            return out

        led = tints(ledger, 2, "Direction")
        rec = tints(recon, 1, "Status")
        assert led, "ledger has no coloured payment rows"
        for k, v in led.items():
            assert len(v) == 1, f"ledger flat {k} has multiple tints {v}"
            if k in rec:
                assert v == rec[k], f"flat {k} tint differs between Ledger and Reconciliation"

    def test_pdf_same_flat_same_tint_on_both_meter_rows(self, client, pid, extra_meter):
        flat, _mid = extra_meter
        pdf = client.get(f"{API}/reports/pack", params={"property_id": pid, "month": MONTH,
                                                        "report": "meters", "format": "pdf"}).content
        doc = fitz.open(stream=pdf, filetype="pdf")
        page = doc[0]
        zoom = 3.0
        pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom))
        stmt_meters = stmt_of(client, pid)["meters"]

        def sample(label):
            hits = page.search_for(label)
            assert hits, f"meter {label} not in PDF"
            r = hits[0]
            return pix.pixel(int((r.x0 - 12) * zoom), int(((r.y0 + r.y1) / 2) * zoom))

        by_flat = {}
        for m in stmt_meters:
            by_flat.setdefault(str(m.get("flat_number")), set()).add(sample(m.get("label")))
        target = str(flat["flat_number"])
        assert len(by_flat[target]) == 1, \
            f"flat {target} meter rows have different tints: {by_flat[target]}"
        uniq = {k: v.pop() for k, v in by_flat.items() if len(v) == 1}
        assert len(uniq) == len(by_flat), f"some flats have mixed tints: {by_flat}"
        assert len(set(uniq.values())) == len(uniq), f"tints collide across flats: {uniq}"


# ------------------------------------------------------------------ (f) layout
class TestOnePageAndNoColourKey:
    @pytest.mark.parametrize("report", ["meters", "purchases", "recurring", "reconciliation"])
    def test_single_report_is_one_page(self, client, pid, report):
        r = client.get(f"{API}/reports/pack", params={"property_id": pid, "month": MONTH,
                                                      "report": report, "format": "pdf"})
        assert r.status_code == 200, r.text
        doc = fitz.open(stream=r.content, filetype="pdf")
        assert doc.page_count == 1, f"{report} report spans {doc.page_count} pages"

    def test_combined_pack_is_cover_plus_four(self, client, pid):
        r = client.get(f"{API}/reports/pack", params={"property_id": pid, "month": MONTH,
                                                      "report": "all", "format": "pdf"})
        assert r.status_code == 200
        doc = fitz.open(stream=r.content, filetype="pdf")
        assert doc.page_count == 5, f"combined pack has {doc.page_count} pages"

    def test_no_colour_key_in_pdfs(self, client, pid):
        for report in ["all", "meters", "purchases", "recurring", "reconciliation"]:
            pdf = client.get(f"{API}/reports/pack", params={"property_id": pid, "month": MONTH,
                                                            "report": report, "format": "pdf"}).content
            text = "".join(p.get_text() for p in fitz.open(stream=pdf, filetype="pdf"))
            assert "Colour key" not in text, f"colour key block still in {report} PDF"

    def test_no_colour_key_in_xlsx(self, client, pid):
        xb = client.get(f"{API}/mis/export", params={"property_id": pid, "month": MONTH,
                                                     "format": "xlsx"}).content
        wb = load_workbook(io.BytesIO(xb))
        offenders = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and "Colour key" in c.value:
                        offenders.append(f"{ws.title}!{c.coordinate}: {c.value}")
        assert not offenders, offenders

    def test_nothing_clipped_horizontally(self, client, pid):
        pdf = client.get(f"{API}/reports/pack", params={"property_id": pid, "month": MONTH,
                                                        "format": "pdf"}).content
        doc = fitz.open(stream=pdf, filetype="pdf")
        problems = []
        for i, page in enumerate(doc):
            limit = page.rect.width - 12 * 72 / 25.4 + 1
            for d in page.get_drawings():
                if d["rect"].x1 > limit:
                    problems.append(f"page {i + 1}: drawing right edge {d['rect'].x1:.1f} > {limit:.1f}")
                    break
            for b in page.get_text("blocks"):
                if b[2] > limit:
                    problems.append(f"page {i + 1}: text right edge {b[2]:.1f} > {limit:.1f}")
                    break
        assert not problems, problems


# ------------------------------------------------------------------ regression
class TestRegression:
    @pytest.mark.parametrize("fmt", ["pdf", "png", "zip"])
    def test_pack_formats(self, client, pid, fmt):
        r = client.get(f"{API}/reports/pack", params={"property_id": pid, "month": MONTH, "format": fmt})
        assert r.status_code == 200, r.text
        assert len(r.content) > 2000

    @pytest.mark.parametrize("fmt", ["csv", "xlsx", "pdf"])
    def test_mis_formats(self, client, pid, fmt):
        r = client.get(f"{API}/mis/export", params={"property_id": pid, "month": MONTH, "format": fmt})
        assert r.status_code == 200, r.text
        assert len(r.content) > 500

    def test_bad_params_rejected(self, client, pid):
        assert client.get(f"{API}/reports/pack", params={"property_id": pid, "month": MONTH,
                                                         "format": "docx"}).status_code == 400
        assert client.get(f"{API}/reports/pack", params={"property_id": pid, "month": MONTH,
                                                         "report": "nope"}).status_code == 400

    def test_flat_ordering_and_data_intact(self, client, pid):
        rows = stmt_of(client, pid)["rows"]
        floors = [r.get("floor") for r in rows]
        order = ["Ground", "First", "Second", "Third"]
        ranks = [order.index(f) if f in order else 99 for f in floors]
        assert ranks == sorted(ranks), floors
        f101 = next((r for r in rows if str(r["flat_number"]) == "101"), None)
        assert f101 and f101["owner_name"] == "Ramesh Kumar", f101

    def test_booking_after_delivery_rejected(self, client, pid):
        r = client.post(f"{API}/tankers", json={
            "property_id": pid, "month": MONTH, "supplier": "TEST_iter14", "date": f"{MONTH}-05",
            "booking_date": f"{MONTH}-09", "qty_sump": 1000, "amount": 500})
        assert r.status_code == 400, r.text
