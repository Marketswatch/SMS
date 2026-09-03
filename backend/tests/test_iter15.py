"""Iteration 15 — pay-back of owner credit, flat-specific charges, grouped readings,
grouped water-usage report, centred report headings, negative opening balance.

Run serially:  python -m pytest tests/test_iter15.py -n 0 -v
"""
import os
import re

import fitz
import pytest
import requests
from dotenv import dotenv_values
from openpyxl import load_workbook
import io

fe = dotenv_values("/app/frontend/.env")
BASE = (os.environ.get("REACT_APP_BACKEND_URL") or fe.get("REACT_APP_BACKEND_URL")).rstrip("/")
API = f"{BASE}/api"
MONTH = "2026-08"


@pytest.fixture(scope="module")
def client():
    s = requests.Session()
    r = s.post(f"{API}/auth/login", json={"email": "admin@societyhub.com", "password": "admin123"})
    assert r.status_code == 200, r.text
    tok = r.json().get("access_token") or r.json().get("token")
    if tok:
        s.headers.update({"Authorization": f"Bearer {tok}"})
    return s


@pytest.fixture(scope="module")
def pid(client):
    props = client.get(f"{API}/properties").json()
    named = [p for p in props if p["name"] == "Sunrise Residency"]
    assert named, "Sunrise Residency property missing"
    return named[0]["id"]


@pytest.fixture(scope="module")
def flats(client, pid):
    return client.get(f"{API}/flats", params={"property_id": pid}).json()


def stmt_of(client, pid):
    r = client.get(f"{API}/statement", params={"property_id": pid, "month": MONTH})
    assert r.status_code == 200, r.text
    return r.json()


def row_of(stmt, number):
    for r in stmt["rows"]:
        if str(r["flat_number"]) == str(number):
            return r
    raise AssertionError(f"flat {number} not in statement")


# ------------------------------------------------------- payment mode / reference
class TestPaymentModeReference:
    def test_non_cash_without_reference_rejected(self, client, pid, flats):
        r = client.post(f"{API}/payments", json={
            "property_id": pid, "month": MONTH, "flat_id": flats[0]["id"],
            "amount": 100, "date": f"{MONTH}-05", "mode": "upi", "reference": "  "})
        assert r.status_code == 400, f"expected 400, got {r.status_code}: {r.text[:300]}"
        assert "reference" in r.text.lower()

    def test_non_cash_with_reference_accepted_and_persisted(self, client, pid, flats):
        fid = flats[0]["id"]
        r = client.post(f"{API}/payments", json={
            "property_id": pid, "month": MONTH, "flat_id": fid, "amount": 100,
            "date": f"{MONTH}-05", "mode": "bank", "reference": "TEST_REF_9911",
            "notes": "TEST_iter15"})
        assert r.status_code == 200, r.text
        pay = r.json()
        assert pay["mode"] == "bank" and pay["reference"] == "TEST_REF_9911"
        assert "_id" not in pay and "id" in pay
        try:
            listed = client.get(f"{API}/payments", params={"property_id": pid, "month": MONTH}).json()
            got = [p for p in listed if p["id"] == pay["id"]]
            assert got and got[0]["reference"] == "TEST_REF_9911" and got[0]["mode"] == "bank"
        finally:
            assert client.delete(f"{API}/payments/{pay['id']}").status_code == 200

    def test_cash_without_reference_ok(self, client, pid, flats):
        r = client.post(f"{API}/payments", json={
            "property_id": pid, "month": MONTH, "flat_id": flats[0]["id"], "amount": 50,
            "date": f"{MONTH}-05", "mode": "cash"})
        assert r.status_code == 200, r.text
        client.delete(f"{API}/payments/{r.json()['id']}")

    def test_payout_non_cash_without_reference_rejected(self, client, pid, flats):
        r = client.post(f"{API}/payments", json={
            "property_id": pid, "month": MONTH, "flat_id": flats[0]["id"], "amount": 100,
            "date": f"{MONTH}-05", "direction": "payout", "mode": "upi", "reference": ""})
        assert r.status_code == 400, f"payout must also require a reference: {r.status_code}"


# ------------------------------------------------------- pay back (partial + full)
class TestPayBackFlow:
    def test_partial_then_full_payback_status(self, client, pid, flats):
        created = []
        try:
            base = stmt_of(client, pid)
            credit_rows = [r for r in base["rows"] if r["payment_status"] == "prepaid"]
            if credit_rows:
                row = credit_rows[0]
                topup = None
            else:
                row = base["rows"][0]
                need = row["net"] + 1000
                rr = client.post(f"{API}/payments", json={
                    "property_id": pid, "month": MONTH, "flat_id": row["flat_id"],
                    "amount": round(need, 2), "date": f"{MONTH}-06", "mode": "cash",
                    "notes": "TEST_iter15 topup"})
                assert rr.status_code == 200, rr.text
                created.append(rr.json()["id"])
                topup = rr.json()
                row = row_of(stmt_of(client, pid), row["flat_number"])
                assert row["payment_status"] == "prepaid", f"topup did not create credit: {row['net']}"
            del topup
            credit = abs(row["net"])
            flat_no = row["flat_number"]
            part = round(credit / 2, 2)

            # part payout with UPI reference
            r1 = client.post(f"{API}/payments", json={
                "property_id": pid, "month": MONTH, "flat_id": row["flat_id"], "amount": part,
                "date": f"{MONTH}-07", "direction": "payout", "mode": "upi",
                "reference": "TEST_UPI_001", "notes": "TEST_iter15 part payback"})
            assert r1.status_code == 200, r1.text
            created.append(r1.json()["id"])

            after = row_of(stmt_of(client, pid), flat_no)
            assert after["payouts"] == pytest.approx(part, abs=0.02)
            assert abs(after["net"]) == pytest.approx(credit - part, abs=0.05), \
                f"credit did not shrink: was {credit}, now {after['net']}"
            assert after["payment_status"] == "prepaid", \
                f"status after PART payback should stay Prepaid, got {after['payment_status']}"

            # remainder
            rest = round(abs(after["net"]), 2)
            r2 = client.post(f"{API}/payments", json={
                "property_id": pid, "month": MONTH, "flat_id": row["flat_id"], "amount": rest,
                "date": f"{MONTH}-08", "direction": "payout", "mode": "cash",
                "notes": "TEST_iter15 final payback"})
            assert r2.status_code == 200, r2.text
            created.append(r2.json()["id"])

            final = row_of(stmt_of(client, pid), flat_no)
            assert abs(final["net"]) <= 0.05, f"net should settle to 0, got {final['net']}"
            assert final["payment_status"] == "excess_paid_back", \
                f"status should be excess_paid_back, got {final['payment_status']}"
        finally:
            for p in created:
                client.delete(f"{API}/payments/{p}")


# ------------------------------------------------------- flat-specific charges
class TestFlatSpecificCharge:
    @pytest.mark.parametrize("ctype,label", [("maintenance", "one-time"), ("cleaning", "recurring")])
    def test_flat_only_charge_maths(self, client, pid, flats, ctype, label):
        target = [f for f in flats if str(f["number"]) == "102"]
        assert target, "flat 102 missing"
        tflat = target[0]
        before = stmt_of(client, pid)
        b102 = row_of(before, "102")
        others_before = {r["flat_number"]: r for r in before["rows"] if r["flat_number"] != "102"}

        cr = client.post(f"{API}/charges", json={
            "property_id": pid, "month": MONTH, "charge_type": ctype,
            "description": f"TEST_iter15 flat-only {label}", "amount": 3000,
            "billed_flat_id": tflat["id"], "date": f"{MONTH}-10"})
        assert cr.status_code == 200, cr.text
        cid = cr.json()["id"]
        assert cr.json().get("billed_flat_id") == tflat["id"]
        try:
            after = stmt_of(client, pid)
            a102 = row_of(after, "102")
            assert a102["flat_specific"] == 3000, f"flat_specific for 102 = {a102['flat_specific']}"
            assert after["totals"]["flat_specific_total"] == 3000
            for r in after["rows"]:
                if r["flat_number"] != "102":
                    assert r["flat_specific"] == 0, f"flat {r['flat_number']} got flat_specific {r['flat_specific']}"
            # shared shares must be untouched
            assert a102["maintenance_share"] == b102["maintenance_share"], "ad-hoc share changed"
            assert a102["recurring_share"] == b102["recurring_share"], "recurring share changed"
            # 102's total rises by exactly 3000, others unchanged
            assert a102["base_cost"] == pytest.approx(b102["base_cost"] + 3000, abs=0.02)
            for r in after["rows"]:
                if r["flat_number"] != "102":
                    assert r["base_cost"] == pytest.approx(
                        others_before[r["flat_number"]]["base_cost"], abs=0.02), \
                        f"flat {r['flat_number']} total changed"
            # detail payload
            det = a102.get("flat_specific_detail") or []
            assert any(d["amount"] == 3000 and d["charge_type"] == ctype for d in det), det

            # charges list marks the flat-only row
            lst = client.get(f"{API}/charges", params={"property_id": pid, "month": MONTH}).json()
            mine = [c for c in lst if c["id"] == cid]
            assert mine and mine[0]["billed_flat_id"] == tflat["id"]

            # CSV export column
            csvr = client.get(f"{API}/mis/export", params={"property_id": pid, "month": MONTH, "format": "csv"})
            assert csvr.status_code == 200
            txt = csvr.text
            assert "Flat-specific" in txt, "CSV missing Flat-specific column"
            recon_lines = [ln for ln in txt.splitlines() if re.match(r"^\d+,102,", ln)]
            assert recon_lines and "3000" in recon_lines[0], recon_lines[:2]

            # XLSX export column
            xr = client.get(f"{API}/mis/export", params={"property_id": pid, "month": MONTH, "format": "xlsx"})
            assert xr.status_code == 200
            wb = load_workbook(io.BytesIO(xr.content))
            joined = []
            for ws in wb.worksheets:
                for rowvals in ws.iter_rows(values_only=True):
                    joined.append([str(v) for v in rowvals if v is not None])
            flat_hdr = [r for r in joined if any("Flat-specific" in c for c in r)]
            assert flat_hdr, "XLSX missing Flat-specific column"
            assert any(any(c in ("3000", "3000.0") for c in r) for r in joined), \
                "XLSX has no 3000 flat-specific value"

            # PDF reconciliation page
            pr = client.get(f"{API}/reports/pack", params={"property_id": pid, "month": MONTH,
                                                           "report": "reconciliation", "format": "pdf"})
            assert pr.status_code == 200
            doc = fitz.open(stream=pr.content, filetype="pdf")
            text = "\n".join(p.get_text() for p in doc)
            assert "specific" in text.lower(), "PDF reconciliation missing Flat-specific column"
            assert "3,000.00" in text, "PDF missing 3,000.00 flat-specific amount"
        finally:
            assert client.delete(f"{API}/charges/{cid}").status_code == 200
        # cleanup verified
        post = stmt_of(client, pid)
        assert post["totals"]["flat_specific_total"] == 0


# ------------------------------------------------------- readings grouping / edit
class TestReadingsGrouped:
    def test_readings_have_floor_flat_owner_and_sorted(self, client, pid):
        rows = client.get(f"{API}/readings", params={"property_id": pid, "month": MONTH}).json()
        assert rows, "no meters"
        for r in rows:
            for k in ("floor", "flat_number", "owner_name", "meter_id", "label"):
                assert k in r, f"missing {k} in readings row"
        order = ["ground", "first", "second", "third"]

        def key(r):
            f = str(r["floor"] or "").lower()
            digits = "".join(ch for ch in str(r["flat_number"]) if ch.isdigit())
            return (order.index(f) if f in order else 99, int(digits or 0), str(r["label"] or ""))

        assert [key(r) for r in rows] == sorted(key(r) for r in rows), \
            f"readings not sorted floor->flat->meter: {[(r['floor'], r['flat_number'], r['label']) for r in rows]}"

    def test_values_land_on_own_meter_and_are_re_editable(self, client, pid):
        orig = client.get(f"{API}/readings", params={"property_id": pid, "month": MONTH}).json()
        assert len(orig) >= 2
        try:
            wanted = {}
            for i, r in enumerate(reversed(orig)):
                wanted[r["meter_id"]] = {"opening": float(r["opening"] or 0),
                                         "closing": float(r["opening"] or 0) + 100 + i * 7}
            payload = {"property_id": pid, "month": MONTH,
                       "readings": [{"meter_id": m, "opening": v["opening"], "closing": v["closing"],
                                     "media": []} for m, v in wanted.items()]}
            resp = client.put(f"{API}/readings", json=payload)
            assert resp.status_code == 200, resp.text
            fresh = {r["meter_id"]: r for r in
                     client.get(f"{API}/readings", params={"property_id": pid, "month": MONTH}).json()}
            for m, v in wanted.items():
                assert fresh[m]["closing"] == pytest.approx(v["closing"]), \
                    f"meter {m} closing mismatch: {fresh[m]['closing']} != {v['closing']}"

            # RE-EDIT one meter and confirm it sticks + statement recomputes
            target = list(wanted.keys())[0]
            new_close = wanted[target]["closing"] + 55
            before = stmt_of(client, pid)
            b = [m for m in before["meters"] if m["meter_id"] == target][0]
            resp2 = client.put(f"{API}/readings", json={
                "property_id": pid, "month": MONTH,
                "readings": [{"meter_id": target, "opening": wanted[target]["opening"],
                              "closing": new_close, "media": []}]})
            assert resp2.status_code == 200, resp2.text
            again = {r["meter_id"]: r for r in
                     client.get(f"{API}/readings", params={"property_id": pid, "month": MONTH}).json()}
            assert again[target]["closing"] == pytest.approx(new_close), "re-edit did not stick"
            for m, v in wanted.items():
                if m != target:
                    assert again[m]["closing"] == pytest.approx(v["closing"]), "other meters disturbed"
            after = stmt_of(client, pid)
            a = [m for m in after["meters"] if m["meter_id"] == target][0]
            assert a["consumption"] == pytest.approx(b["consumption"] + 55, abs=0.02), \
                "statement did not recompute after re-edit"
        finally:
            client.put(f"{API}/readings", json={
                "property_id": pid, "month": MONTH,
                "readings": [{"meter_id": r["meter_id"], "opening": r["opening"],
                              "closing": r["closing"], "media": r.get("media", [])} for r in orig]})


# ------------------------------------------------------- grouped water usage report
class TestGroupedMeters:
    def test_two_meters_of_same_flat_group_with_single_total(self, client, pid, flats):
        f101 = [f for f in flats if str(f["number"]) == "101"][0]
        mr = client.post(f"{API}/meters", json={"property_id": pid, "flat_id": f101["id"],
                                                "label": "TEST-M101B", "opening": 0})
        assert mr.status_code == 200, mr.text
        mid = mr.json()["id"]
        orig = client.get(f"{API}/readings", params={"property_id": pid, "month": MONTH}).json()
        try:
            client.put(f"{API}/readings", json={"property_id": pid, "month": MONTH,
                                                "readings": [{"meter_id": mid, "opening": 0,
                                                              "closing": 250, "media": []}]})
            stmt = stmt_of(client, pid)
            m101 = [m for m in stmt["meters"] if str(m["flat_number"]) == "101"]
            assert len(m101) >= 2, "second meter missing from statement"
            idx = [i for i, m in enumerate(stmt["meters"]) if str(m["flat_number"]) == "101"]
            assert idx == list(range(idx[0], idx[0] + len(idx))), \
                f"flat 101 meters not adjacent: {idx}"
            combined = round(sum(float(m["charge"]) for m in m101), 2)

            pr = client.get(f"{API}/reports/pack", params={"property_id": pid, "month": MONTH,
                                                           "report": "meters", "format": "pdf"})
            assert pr.status_code == 200
            doc = fitz.open(stream=pr.content, filetype="pdf")
            assert doc.page_count == 1, f"meters report should be 1 page, got {doc.page_count}"
            text = doc[0].get_text()
            assert "Total\nAmount" in text or "Total Amount" in text or "Amount" in text
            if combined > 0:
                assert f"{combined:,.2f}" in text, f"combined total {combined} not in meters PDF"
            # 'House' column value 101 should appear once in the flat block
            assert text.count("TEST-M101B") == 1
        finally:
            client.delete(f"{API}/meters/{mid}")
            client.put(f"{API}/readings", json={
                "property_id": pid, "month": MONTH,
                "readings": [{"meter_id": r["meter_id"], "opening": r["opening"],
                              "closing": r["closing"], "media": r.get("media", [])} for r in orig]})
            left = client.get(f"{API}/meters", params={"property_id": pid}).json()
            assert not [m for m in left if m["label"] == "TEST-M101B"], "test meter not cleaned up"


# ------------------------------------------------------- report headings / layout
class TestReportHeadings:
    @pytest.mark.parametrize("which", ["meters", "purchases", "recurring", "reconciliation"])
    def test_centred_two_line_heading_no_sort_note(self, client, pid, which):
        r = client.get(f"{API}/reports/pack", params={"property_id": pid, "month": MONTH,
                                                      "report": which, "format": "pdf"})
        assert r.status_code == 200, r.text
        doc = fitz.open(stream=r.content, filetype="pdf")
        assert doc.page_count == 1, f"{which} should be 1 page, got {doc.page_count}"
        page = doc[0]
        text = page.get_text()
        assert "For the month of August 2026" in text, f"{which} missing month line"
        low = text.lower()
        for bad in ("sorted by floor", "then flat number", "sorted by"):
            assert bad not in low, f"{which} still contains '{bad}'"
        # centring: the two heading lines' mid-x should be near the page centre
        mid = page.rect.width / 2
        blocks = [b for b in page.get_text("blocks") if "For the month of" in b[4]
                  or "Sunrise Residency" in b[4]]
        assert blocks, "heading blocks not found"
        for b in blocks:
            bmid = (b[0] + b[2]) / 2
            assert abs(bmid - mid) < 60, f"{which} heading not centred (mid {bmid:.0f} vs {mid:.0f})"

    def test_combined_pack_is_cover_plus_four(self, client, pid):
        r = client.get(f"{API}/reports/pack", params={"property_id": pid, "month": MONTH,
                                                      "report": "all", "format": "pdf"})
        assert r.status_code == 200
        doc = fitz.open(stream=r.content, filetype="pdf")
        assert doc.page_count == 5, f"expected cover + 4 pages, got {doc.page_count}"
        text = "\n".join(p.get_text() for p in doc)
        assert "Water Charges" in text, "missing Water Charges group band"
        assert "sorted by" not in text.lower()

    def test_png_and_zip_still_work(self, client, pid):
        png = client.get(f"{API}/reports/pack", params={"property_id": pid, "month": MONTH,
                                                        "report": "meters", "format": "png"})
        assert png.status_code == 200 and png.content[:4] == b"\x89PNG"
        z = client.get(f"{API}/reports/pack", params={"property_id": pid, "month": MONTH,
                                                      "report": "all", "format": "zip"})
        assert z.status_code == 200 and z.content[:2] == b"PK"


# ------------------------------------------------------- negative opening balance
class TestNegativeOpeningBalance:
    def test_negative_opening_is_payable_to_flat(self, client, pid, flats):
        f = [x for x in flats if str(x["number"]) == "103"] or [flats[-1]]
        flat = f[0]
        before = row_of(stmt_of(client, pid), flat["number"])
        payload = {k: v for k, v in flat.items() if k != "id"}
        payload["opening_dues"] = -2000
        up = client.put(f"{API}/flats/{flat['id']}", json=payload)
        assert up.status_code == 200, up.text
        try:
            after = row_of(stmt_of(client, pid), flat["number"])
            if before["carry_in"] != before["opening_dues"]:
                pytest.skip("month already closed — carry_in comes from previous period")
            assert after["carry_in"] == -2000, f"carry brought forward = {after['carry_in']}"
            assert after["net"] == pytest.approx(before["net"] - 2000 + before["carry_in"] * 0, abs=0.02) \
                or after["net"] < before["net"], "net did not move toward credit"
            if after["net"] < -0.005:
                assert after["payment_status"] == "prepaid", after["payment_status"]
        finally:
            payload["opening_dues"] = 0
            client.put(f"{API}/flats/{flat['id']}", json=payload)
            reset = row_of(stmt_of(client, pid), flat["number"])
            assert reset["opening_dues"] == 0


# ------------------------------------------------------- light regression
class TestRegression:
    def test_statement_annual_overview(self, client, pid):
        for path, params in (("/statement", {"property_id": pid, "month": MONTH}),
                             ("/annual", {"property_id": pid, "year": "2026"}),
                             ("/overview", {"month": MONTH})):
            r = client.get(f"{API}{path}", params=params)
            assert r.status_code == 200, f"{path} -> {r.status_code} {r.text[:200]}"
            assert '"_id"' not in r.text

    def test_mis_pdf_and_annual_export(self, client, pid):
        r = client.get(f"{API}/mis/export", params={"property_id": pid, "month": MONTH, "format": "pdf"})
        assert r.status_code == 200 and r.content[:4] == b"%PDF"
        a = client.get(f"{API}/annual/export", params={"property_id": pid, "year": "2026", "format": "csv"})
        assert a.status_code == 200

    def test_booking_after_delivery_rejected(self, client, pid, flats):
        r = client.post(f"{API}/tankers", json={
            "property_id": pid, "month": MONTH, "supplier": "TEST_iter15",
            "booking_date": f"{MONTH}-20", "date": f"{MONTH}-10",
            "qty_sump": 1000, "qty_syntex": 0, "amount": 500,
            "payer_flat_id": flats[0]["id"], "payer_type": "owner"})
        assert r.status_code == 400, f"expected 400 for booking after delivery, got {r.status_code}"
        if r.status_code == 200:
            client.delete(f"{API}/tankers/{r.json()['id']}")
