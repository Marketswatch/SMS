"""Iteration 11 — styled .xlsx month packs for Maintenance (MIS) and Property/Rentals."""
import io
import os
import re

import pytest
import requests
from dotenv import dotenv_values
from openpyxl import load_workbook

fe = dotenv_values("/app/frontend/.env")
BASE = (os.environ.get("REACT_APP_BACKEND_URL") or fe.get("REACT_APP_BACKEND_URL")).rstrip("/")
API = f"{BASE}/api"
XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
NAVY = "1F3864"
TOTAL_FILL = "C9D7EE"

RECON_HEADERS = ["S.No", "Flat No.", "Floor", "Owner", "Metered cost", "Non-metered cost (reserve)",
                 "Total water cost", "Misc", "Total amount", "Bal brought forward",
                 "Advance paid (fronting)", "Amount paid", "Balance to pay / receive",
                 "Date of payment", "Status"]
MIS_SHEETS = ["Water Reconciliation", "Water Usage (Meters)", "Tanker Purchases", "Charges", "Ledger"]
RENT_SHEETS = ["Rent Roll", "Collections", "Payouts", "Building Settlement", "Deposits"]
DMY = re.compile(r"^\d{2}-\d{2}-\d{4}$")


# ------------------------------------------------------------------ fixtures
@pytest.fixture(scope="session")
def client():
    s = requests.Session()
    r = s.post(f"{API}/auth/login", json={"email": "admin@societyhub.com", "password": "admin123"})
    assert r.status_code == 200, r.text
    tok = r.json().get("access_token") or r.json().get("token")
    if tok:
        s.headers.update({"Authorization": f"Bearer {tok}"})
    return s


@pytest.fixture(scope="session")
def prop(client):
    r = client.get(f"{API}/properties")
    assert r.status_code == 200, r.text
    props = [p for p in r.json() if p.get("kind", "maintenance") != "rental"]
    assert props, "no maintenance property"
    named = [p for p in props if p["name"] == "Sunrise Residency"]
    return (named or props)[0]


@pytest.fixture(scope="session")
def month(client, prop):
    """Prefer the seeded month that actually has tankers/payments (2026-08)."""
    r = client.get(f"{API}/periods", params={"property_id": prop["id"]})
    assert r.status_code == 200, r.text
    months = [p["month"] for p in r.json()]
    return "2026-08" if "2026-08" in months else months[-1]


@pytest.fixture(scope="session")
def statement(client, prop, month):
    r = client.get(f"{API}/statement", params={"property_id": prop["id"], "month": month})
    assert r.status_code == 200, r.text
    return r.json()


@pytest.fixture(scope="session")
def mis_resp(client, prop, month):
    return client.get(f"{API}/mis/export",
                      params={"property_id": prop["id"], "month": month, "format": "xlsx"})


@pytest.fixture(scope="session")
def wb(mis_resp):
    assert mis_resp.status_code == 200, mis_resp.text[:400]
    return load_workbook(io.BytesIO(mis_resp.content))


def find_header_row(ws, first_header):
    for r in range(1, min(ws.max_row, 40) + 1):
        if str(ws.cell(row=r, column=1).value or "").strip() == first_header:
            return r
    return None


def rgb(cell):
    f = cell.fill
    if not f or f.fill_type != "solid":
        return None
    v = f.fgColor.rgb
    return v[-6:] if isinstance(v, str) else None


def cells_text(ws):
    out = []
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                out.append(str(c.value))
    return out


# ------------------------------------------------------------------ MIS pack
class TestMisWorkbook:
    def test_content_type_and_filename(self, mis_resp, prop, month):
        assert mis_resp.status_code == 200
        assert mis_resp.headers["content-type"].split(";")[0] == XLSX_MEDIA
        cd = mis_resp.headers.get("content-disposition", "")
        assert f"{prop['name'].replace(' ', '-').lower()}-{month}.xlsx" in cd, cd
        assert mis_resp.content[:2] == b"PK"

    def test_exact_five_sheets(self, wb):
        assert wb.sheetnames == MIS_SHEETS

    def test_recon_headers_order(self, wb):
        ws = wb["Water Reconciliation"]
        hr = find_header_row(ws, "S.No")
        assert hr, "header row not found"
        got = [str(ws.cell(row=hr, column=j).value) for j in range(1, 16)]
        assert got == RECON_HEADERS
        assert ws.cell(row=hr, column=16).value is None

    def test_header_style_all_sheets(self, wb):
        for name in MIS_SHEETS:
            ws = wb[name]
            hr = find_header_row(ws, "S.No")
            assert hr, f"{name}: no header row"
            c = ws.cell(row=hr, column=1)
            assert rgb(c) == NAVY, f"{name}: header fill {rgb(c)}"
            assert c.font.bold and c.font.color.rgb[-6:] == "FFFFFF", f"{name}: header font"

    def test_frozen_panes_and_widths(self, wb):
        for name in ["Water Reconciliation", "Water Usage (Meters)", "Tanker Purchases", "Ledger"]:
            ws = wb[name]
            assert ws.freeze_panes, f"{name}: no frozen panes"
            assert ws.column_dimensions["A"].width and ws.column_dimensions["A"].width >= 5
            assert ws.column_dimensions["D"].width and ws.column_dimensions["D"].width >= 10

    def test_recon_values_match_statement(self, wb, statement):
        ws = wb["Water Reconciliation"]
        hr = find_header_row(ws, "S.No")
        rows = statement["rows"]
        for i, r in enumerate(rows):
            row = hr + 1 + i
            vals = [ws.cell(row=row, column=j).value for j in range(1, 16)]
            assert vals[0] == i + 1
            assert vals[1] == r["flat_number"]
            assert vals[3] == r["owner_name"], f"owner mismatch {vals[3]}"
            assert round(float(vals[4]), 2) == round(r["water_own_cost"], 2)
            assert round(float(vals[5]), 2) == round(r["reserve_share"], 2)
            assert round(float(vals[6]), 2) == round(r["water_cost"], 2)
            assert round(float(vals[8]), 2) == round(r["base_cost"], 2)
            assert round(float(vals[9]), 2) == round(r["carry_in"], 2)
            assert round(float(vals[10]), 2) == round(r["contributions"], 2)
            assert round(float(vals[11]), 2) == round(r["received"], 2)
            assert round(float(vals[12]), 2) == round(r["net"], 2)
            if vals[13] and str(vals[13]) != "—":
                assert DMY.match(str(vals[13])), vals[13]
            # money formats + borders on data cells
            for j in range(5, 14):
                c = ws.cell(row=row, column=j)
                assert c.number_format == "#,##0.00", f"col {j} format {c.number_format}"
                assert c.border.left.style == "thin"

    def test_recon_total_row_shaded(self, wb, statement):
        ws = wb["Water Reconciliation"]
        hr = find_header_row(ws, "S.No")
        trow = hr + 1 + len(statement["rows"])
        assert ws.cell(row=trow, column=4).value == "TOTAL"
        assert rgb(ws.cell(row=trow, column=4)) == TOTAL_FILL
        t = statement["totals"]
        assert round(float(ws.cell(row=trow, column=9).value), 2) == round(t["billable_total"], 2)
        assert round(float(ws.cell(row=trow, column=13).value), 2) == round(t["net_position"], 2)
        assert ws.cell(row=trow, column=9).number_format == "#,##0.00"

    def test_balance_red_green(self, wb, statement):
        ws = wb["Water Reconciliation"]
        hr = find_header_row(ws, "S.No")
        seen_pos = seen_neg = False
        for i, r in enumerate(statement["rows"]):
            c = ws.cell(row=hr + 1 + i, column=13)
            col = (c.font.color.rgb or "")[-6:] if c.font.color else ""
            if r["net"] > 0:
                assert c.font.bold and col == "C00000", f"{r['flat_number']} net>0 colour {col}"
                seen_pos = True
            elif r["net"] < 0:
                assert c.font.bold and col == "1F7A3D", f"{r['flat_number']} net<0 colour {col}"
                seen_neg = True
        assert seen_pos or seen_neg, "no signed balances to verify"

    def test_owner_colour_coding_and_key(self, wb, statement):
        ws = wb["Water Reconciliation"]
        hr = find_header_row(ws, "S.No")
        by_owner = {}
        for i, r in enumerate(statement["rows"]):
            fills = {rgb(ws.cell(row=hr + 1 + i, column=j)) for j in range(1, 16)}
            assert len(fills) == 1, f"row {i} has mixed fills {fills}"
            tint = fills.pop()
            assert tint, f"row {i} has no tint"
            by_owner.setdefault(r["owner_name"], set()).add(tint)
        for owner, tints in by_owner.items():
            assert len(tints) == 1, f"{owner} has inconsistent tints {tints}"
        distinct = {list(v)[0] for v in by_owner.values()}
        assert len(distinct) == len(by_owner), f"owners share tints: {by_owner}"
        text = cells_text(ws)
        assert "Colour key — owner" in text
        for owner in by_owner:
            assert owner in text, f"{owner} missing from colour key"

    def test_summary_legend_block(self, wb, statement):
        ws = wb["Water Reconciliation"]
        text = cells_text(ws)
        assert "Summary" in text
        for label in ["Total expense for the month", "Expense per head",
                      "Total receivable (owes)", "Total payable (owed to owners)", "Net position"]:
            assert label in text, f"missing legend row: {label}"
        assert any(l.startswith("Split between") for l in text)

    def test_original_owner_names_no_tenant_suffix(self, wb, statement):
        ws = wb["Water Reconciliation"]
        hr = find_header_row(ws, "S.No")
        for i, r in enumerate(statement["rows"]):
            v = ws.cell(row=hr + 1 + i, column=4).value
            assert v == r["owner_name"]
            assert "tenant" not in str(v).lower() and "(" not in str(v)

    # ------------------------------------------------- meters sheet
    def test_meter_sheet_colour_per_meter(self, wb, statement):
        ws = wb["Water Usage (Meters)"]
        hr = find_header_row(ws, "S.No")
        heads = [str(ws.cell(row=hr, column=j).value) for j in range(1, 11)]
        assert heads == ["S.No", "House", "Floor", "Owner", "Meter number", "Starting unit",
                         "Ending unit", "Consumed units", "Water charges", "Combined (per flat)"]
        meters = statement["meters"]
        tint_by_meter, flat_tints = {}, {}
        for i, m in enumerate(meters):
            row = hr + 1 + i
            assert ws.cell(row=row, column=5).value == m.get("label")
            tint = rgb(ws.cell(row=row, column=5))
            assert tint, f"meter {m.get('label')} row has no tint"
            tint_by_meter.setdefault(m["label"], set()).add(tint)
            flat_tints.setdefault(m.get("flat_number"), set()).add(tint)
        assert len({list(v)[0] for v in tint_by_meter.values()}) == len(tint_by_meter), \
            f"meters share tints: {tint_by_meter}"
        multi = {k: v for k, v in flat_tints.items() if len(
            [m for m in meters if m.get("flat_number") == k]) > 1}
        for flat, tints in multi.items():
            assert len(tints) > 1, f"flat {flat} with 2 meters has one tint {tints}"
        text = cells_text(ws)
        assert "Colour key — meter" in text
        for label in tint_by_meter:
            assert label in text

    def test_meter_total_row(self, wb, statement):
        ws = wb["Water Usage (Meters)"]
        hr = find_header_row(ws, "S.No")
        trow = hr + 1 + len(statement["meters"])
        assert ws.cell(row=trow, column=5).value == "TOTAL"
        assert rgb(ws.cell(row=trow, column=5)) == TOTAL_FILL
        assert round(float(ws.cell(row=trow, column=9).value), 2) == \
            round(statement["totals"]["metered_charges"], 2)

    # ------------------------------------------------- tankers sheet
    def test_tanker_sheet(self, wb, client, prop, month):
        ws = wb["Tanker Purchases"]
        hr = find_header_row(ws, "S.No")
        heads = [str(ws.cell(row=hr, column=j).value) for j in range(1, 13)]
        assert heads == ["S.No", "Date", "Supplier", "Sump (L)", "Syntex (L)", "Total (L)",
                         "Lorry amount", "Tips", "Total cost", "Cost / L",
                         "Lorry paid by", "Tips paid by"]
        tk = client.get(f"{API}/tankers", params={"property_id": prop["id"], "month": month}).json()
        assert len(tk) > 0
        diff_payer = False
        for i in range(len(tk)):
            row = hr + 1 + i
            assert ws.cell(row=row, column=1).value == i + 1
            assert DMY.match(str(ws.cell(row=row, column=2).value))
            lorry = str(ws.cell(row=row, column=11).value)
            tips = str(ws.cell(row=row, column=12).value)
            if tips != "—" and tips.split(" (")[0] != lorry.split(" (")[0]:
                diff_payer = True
        assert diff_payer, "no tanker where tips payer differs from lorry payer"
        trow = hr + 1 + len(tk)
        assert ws.cell(row=trow, column=3).value == "TOTAL"
        assert rgb(ws.cell(row=trow, column=3)) == TOTAL_FILL
        text = cells_text(ws)
        assert "Summary" in text and "Expense per head" in text

    # ------------------------------------------------- charges sheet
    def test_charges_sections(self, wb):
        ws = wb["Charges"]
        text = cells_text(ws)
        bands = [t for t in text if t.startswith("Recurring —") or t.startswith("One-time / repairs —")]
        assert len(bands) == 2, f"section bands: {bands}"
        for t in bands:
            r = next(r for r in range(1, ws.max_row + 1)
                     if str(ws.cell(row=r, column=1).value or "") == t)
            assert rgb(ws.cell(row=r, column=1)) == "2F5597"
        totals = [r for r in range(1, ws.max_row + 1)
                  if str(ws.cell(row=r, column=3).value or "") == "TOTAL"]
        assert len(totals) == 2, f"expected 2 TOTAL rows, got {len(totals)}"
        for r in totals:
            assert rgb(ws.cell(row=r, column=3)) == TOTAL_FILL
            assert str(ws.cell(row=r, column=8).value or "").startswith("Per head")

    # ------------------------------------------------- ledger sheet
    def test_ledger_sheet(self, wb, client, prop, month, statement):
        ws = wb["Ledger"]
        hr = find_header_row(ws, "S.No")
        heads = [str(ws.cell(row=hr, column=j).value) for j in range(1, 9)]
        assert heads == ["S.No", "Date", "Flat No.", "Owner", "Direction", "Paid by", "Amount", "Notes"]
        pays = client.get(f"{API}/payments", params={"property_id": prop["id"], "month": month}).json()
        assert len(pays) > 0
        owner_tint = {}
        for i in range(len(pays)):
            row = hr + 1 + i
            d = ws.cell(row=row, column=5).value
            assert d in ("Received", "Payout"), d
            assert DMY.match(str(ws.cell(row=row, column=2).value))
            assert ws.cell(row=row, column=7).number_format == "#,##0.00"
            owner = ws.cell(row=row, column=4).value
            if owner:
                owner_tint.setdefault(owner, set()).add(rgb(ws.cell(row=row, column=4)))
        for owner, tints in owner_tint.items():
            assert len(tints) == 1 and list(tints)[0], f"{owner} tint {tints}"
        trow = hr + 1 + len(pays)
        assert ws.cell(row=trow, column=4).value == "TOTAL"
        assert round(float(ws.cell(row=trow, column=7).value), 2) == \
            round(statement["totals"]["total_received"], 2)
        assert "Colour key — owner" in cells_text(ws)

    def test_ledger_colour_key_lists_only_owners(self, wb, statement):
        """The shared owner palette also receives flat numbers (tanker/charge payers),
        so the Ledger 'Colour key — owner' block leaks flat numbers as owner entries."""
        ws = wb["Ledger"]
        start = next(r for r in range(1, ws.max_row + 1)
                     if str(ws.cell(row=r, column=1).value or "") == "Colour key — owner")
        keys = [str(ws.cell(row=r, column=1).value) for r in range(start + 1, ws.max_row + 1)
                if ws.cell(row=r, column=1).value]
        owners = {r["owner_name"] for r in statement["rows"]}
        assert set(keys) == owners, f"colour key contains non-owner entries: {set(keys) - owners}"

    def test_owner_tints_consistent_across_sheets(self, wb, statement):
        """Same owner must keep the same tint on Reconciliation and Ledger."""
        ws = wb["Water Reconciliation"]
        hr = find_header_row(ws, "S.No")
        recon = {r["owner_name"]: rgb(ws.cell(row=hr + 1 + i, column=4))
                 for i, r in enumerate(statement["rows"])}
        ws5 = wb["Ledger"]
        hr5 = find_header_row(ws5, "S.No")
        for row in range(hr5 + 1, ws5.max_row + 1):
            owner = ws5.cell(row=row, column=4).value
            if owner in recon:
                assert rgb(ws5.cell(row=row, column=4)) == recon[owner], \
                    f"{owner} tint differs across sheets"


# ------------------------------------------------------------------ rename verbatim
class TestVerbatimRename:
    def test_owner_rename_reflected(self, client, prop, month, statement):
        target = statement["rows"][0]
        flat_id = target["flat_id"]
        flats = client.get(f"{API}/flats", params={"property_id": prop["id"]}).json()
        doc = next(f for f in flats if f["id"] == flat_id)
        body = {k: v for k, v in doc.items() if k not in ("id", "created_at")}
        original = body.get("owner_name")
        new_name = "QA Verbatim O'Brien-Kumar  Jr."
        r = client.put(f"{API}/flats/{flat_id}", json={**body, "owner_name": new_name})
        assert r.status_code == 200, r.text
        try:
            resp = client.get(f"{API}/mis/export", params={"property_id": prop["id"],
                                                          "month": month, "format": "xlsx"})
            assert resp.status_code == 200
            ws = load_workbook(io.BytesIO(resp.content))["Water Reconciliation"]
            hr = find_header_row(ws, "S.No")
            names = [ws.cell(row=hr + 1 + i, column=4).value for i in range(len(statement["rows"]))]
            assert new_name in names, names
        finally:
            client.put(f"{API}/flats/{flat_id}", json={**body, "owner_name": original})

    def test_meter_rename_not_supported_by_api(self, client, prop, statement):
        """There is no PUT/PATCH /api/meters/{id} — a meter cannot be renamed (only add/delete)."""
        m = statement["meters"][0]
        mid = m.get("meter_id") or m.get("id")
        r = client.put(f"{API}/meters/{mid}", json={"label": "QA Meter 9X/1"})
        assert r.status_code == 405, f"meter rename now supported? {r.status_code}"

    def test_new_meter_label_verbatim(self, client, prop, month, statement):
        """Special-character meter label must appear verbatim in the Excel meter column."""
        flat_id = statement["rows"][0]["flat_id"]
        label = "QA Meter 9X/1 (Sump) #2"
        r = client.post(f"{API}/meters", json={"property_id": prop["id"], "flat_id": flat_id,
                                              "label": label, "opening": 10})
        assert r.status_code == 200, r.text
        mid = r.json()["id"]
        try:
            resp = client.get(f"{API}/mis/export", params={"property_id": prop["id"],
                                                          "month": month, "format": "xlsx"})
            assert resp.status_code == 200, resp.text[:200]
            ws = load_workbook(io.BytesIO(resp.content))["Water Usage (Meters)"]
            assert label in cells_text(ws), "new meter label missing from Excel"
            # a flat with TWO meters must show TWO different tints
            hr = find_header_row(ws, "S.No")
            flat_no = statement["rows"][0]["flat_number"]
            tints = [rgb(ws.cell(row=r, column=5)) for r in range(hr + 1, ws.max_row + 1)
                     if str(ws.cell(row=r, column=2).value) == str(flat_no)
                     and ws.cell(row=r, column=5).value not in (None, "TOTAL")]
            assert len(tints) >= 2, f"expected 2 meter rows for flat {flat_no}, got {tints}"
            assert len(set(tints)) == len(tints), f"two meters of flat {flat_no} share a tint {tints}"
        finally:
            client.delete(f"{API}/meters/{mid}")


# ------------------------------------------------------------------ rentals pack
@pytest.fixture(scope="session")
def rent_month(client):
    pays = client.get(f"{API}/rentals/payments").json()
    months = sorted({p["month"] for p in pays})
    return months[-1] if months else "2026-08"


@pytest.fixture(scope="session")
def rent_seed(client, rent_month):
    """Seed a 2nd collection (cash + bank modes), a payout and a deposit so every
    rentals sheet has data; everything is removed afterwards."""
    units = client.get(f"{API}/rentals/units").json()
    created = {"payments": [], "payouts": [], "deposits": []}
    if len(units) >= 2:
        for u, mode, ref in ((units[0], "cash", ""), (units[1], "bank", "TEST-NEFT-9001")):
            r = client.post(f"{API}/rentals/payments", json={
                "unit_id": u["id"], "month": rent_month, "date": f"{rent_month}-12",
                "rent_paid": 1000, "maintenance_paid": 100, "adhoc_paid": 0,
                "mode": mode, "reference": ref, "notes": "TEST_xlsx"})
            if r.status_code == 200:
                created["payments"].append(r.json()["id"])
        r = client.post(f"{API}/rentals/payouts", json={
            "building_name": units[0].get("building_name") or "TEST Building",
            "unit_id": units[0]["id"], "month": rent_month, "amount": 500,
            "date": f"{rent_month}-14", "category": "Maintenance", "note": "TEST_xlsx",
            "mode": "upi", "reference": "TEST-UPI-77"})
        if r.status_code == 200:
            created["payouts"].append(r.json()["id"])
        r = client.post(f"{API}/rentals/deposits", json={
            "unit_id": units[0]["id"], "month": rent_month, "kind": "deposit",
            "amount": 5000, "date": f"{rent_month}-02", "mode": "bank", "notes": "TEST_xlsx"})
        if r.status_code == 200:
            created["deposits"].append(r.json()["id"])
    yield created
    for kind, ids in created.items():
        for i in ids:
            client.delete(f"{API}/rentals/{kind}/{i}")


@pytest.fixture(scope="session")
def rent_data(client, rent_month, rent_seed):
    r = client.get(f"{API}/rentals/statement", params={"month": rent_month})
    assert r.status_code == 200, r.text
    return r.json()


@pytest.fixture(scope="session")
def rent_resp(client, rent_month, rent_seed):
    return client.get(f"{API}/rentals/export", params={"month": rent_month, "format": "xlsx"})


@pytest.fixture(scope="session")
def rwb(rent_resp):
    assert rent_resp.status_code == 200, rent_resp.text[:400]
    return load_workbook(io.BytesIO(rent_resp.content))


class TestRentalsWorkbook:
    def test_media_type_and_sheets(self, rent_resp, rwb, rent_month):
        assert rent_resp.headers["content-type"].split(";")[0] == XLSX_MEDIA
        assert f"societyhub-properties-{rent_month}.xlsx" in \
            rent_resp.headers.get("content-disposition", "")
        assert rwb.sheetnames == RENT_SHEETS

    def test_headers_styled(self, rwb):
        for name in RENT_SHEETS:
            ws = rwb[name]
            hr = find_header_row(ws, "S.No")
            assert hr, f"{name}: no header"
            c = ws.cell(row=hr, column=1)
            assert rgb(c) == NAVY and c.font.bold and c.font.color.rgb[-6:] == "FFFFFF", name

    def test_rent_roll_values_and_tints(self, rwb, rent_data):
        ws = rwb["Rent Roll"]
        hr = find_header_row(ws, "S.No")
        heads = [str(ws.cell(row=hr, column=j).value) for j in range(1, 18)]
        assert heads[:5] == ["S.No", "Property", "Building", "Tenant", "Rent"]
        assert heads[14:16] == ["Balance", "Status"]
        tints = {}
        for i, r in enumerate(rent_data["rows"]):
            row = hr + 1 + i
            assert ws.cell(row=row, column=2).value == r["name"]
            assert round(float(ws.cell(row=row, column=10).value), 2) == round(r["total_to_collect"], 2)
            assert round(float(ws.cell(row=row, column=15).value), 2) == round(r["balance"], 2)
            assert ws.cell(row=row, column=5).number_format == "#,##0.00"
            assert ws.cell(row=row, column=5).border.left.style == "thin"
            t = rgb(ws.cell(row=row, column=2))
            assert t, f"{r['name']} has no tint"
            tints.setdefault(r["name"], set()).add(t)
        assert len({list(v)[0] for v in tints.values()}) == len(tints), f"shared tints {tints}"
        text = cells_text(ws)
        assert "Summary" in text and "Colour key — property" in text
        for label in ["Properties", "Occupied", "Vacant", "To collect", "Collected",
                      "Balance outstanding", "Deposits held"]:
            assert label in text, label
        trow = hr + 1 + len(rent_data["rows"])
        assert ws.cell(row=trow, column=2).value == "TOTAL"
        assert rgb(ws.cell(row=trow, column=2)) == TOTAL_FILL
        assert ws.freeze_panes

    def test_collections_modes_and_reference(self, rwb, client, rent_month):
        ws = rwb["Collections"]
        hr = find_header_row(ws, "S.No")
        heads = [str(ws.cell(row=hr, column=j).value) for j in range(1, 11)]
        assert heads == ["S.No", "Date", "Property", "Rent", "Maintenance", "Ad-hoc", "Total",
                         "Mode", "Reference", "Notes"]
        pays = client.get(f"{API}/rentals/payments", params={"month": rent_month}).json()
        assert len(pays) > 0, "no rental collections seeded"
        modes = set()
        for i in range(len(pays)):
            row = hr + 1 + i
            assert DMY.match(str(ws.cell(row=row, column=2).value)), ws.cell(row=row, column=2).value
            modes.add(str(ws.cell(row=row, column=8).value))
            assert ws.cell(row=row, column=9).value not in (None, "")
        assert modes <= {"Cash", "UPI", "Bank Transfer", "Cheque", "Card"}, modes
        trow = hr + 1 + len(pays)
        assert ws.cell(row=trow, column=3).value == "TOTAL"

    def test_payouts_and_settlement_and_deposits(self, rwb, rent_data):
        ws = rwb["Payouts"]
        hr = find_header_row(ws, "S.No")
        assert [str(ws.cell(row=hr, column=j).value) for j in range(1, 11)] == \
            ["S.No", "Date", "Building", "Property", "Category", "Amount", "Mode",
             "Reference", "Type", "Note"]
        ws4 = rwb["Building Settlement"]
        hr4 = find_header_row(ws4, "S.No")
        for i, b in enumerate(rent_data["buildings"]):
            row = hr4 + 1 + i
            assert ws4.cell(row=row, column=2).value == b["building"]
            assert round(float(ws4.cell(row=row, column=6).value), 2) == round(b["balance"], 2)
        trow = hr4 + 1 + len(rent_data["buildings"])
        assert ws4.cell(row=trow, column=2).value == "TOTAL"
        ws5 = rwb["Deposits"]
        hr5 = find_header_row(ws5, "S.No")
        assert [str(ws5.cell(row=hr5, column=j).value) for j in range(1, 7)] == \
            ["S.No", "Date", "Property", "Type", "Amount", "Notes"]

    def test_property_tints_consistent_across_sheets(self, rwb, rent_data):
        ws = rwb["Rent Roll"]
        hr = find_header_row(ws, "S.No")
        base = {r["name"]: rgb(ws.cell(row=hr + 1 + i, column=2))
                for i, r in enumerate(rent_data["rows"])}
        for name, col in (("Collections", 3), ("Payouts", 4), ("Deposits", 3)):
            ws2 = rwb[name]
            h = find_header_row(ws2, "S.No")
            for row in range(h + 1, ws2.max_row + 1):
                v = ws2.cell(row=row, column=col).value
                if v in base:
                    assert rgb(ws2.cell(row=row, column=col)) == base[v], \
                        f"{name}: {v} tint differs"


# ------------------------------------------------------------------ regression
class TestExportRegression:
    def test_mis_csv_pdf(self, client, prop, month):
        c = client.get(f"{API}/mis/export", params={"property_id": prop["id"], "month": month, "format": "csv"})
        assert c.status_code == 200 and "S.No,Flat No.,Floor,Owner" in c.text
        p = client.get(f"{API}/mis/export", params={"property_id": prop["id"], "month": month, "format": "pdf"})
        assert p.status_code == 200 and p.content[:4] == b"%PDF"

    def test_annual_exports(self, client, prop):
        for fmt, sig in (("csv", None), ("pdf", b"%PDF")):
            r = client.get(f"{API}/annual/export",
                           params={"property_id": prop["id"], "year": 2026, "format": fmt})
            assert r.status_code == 200, r.text[:200]
            if sig:
                assert r.content[:4] == sig

    def test_rentals_csv_pdf(self, client, rent_month):
        c = client.get(f"{API}/rentals/export", params={"month": rent_month, "format": "csv"})
        assert c.status_code == 200 and len(c.text) > 50
        p = client.get(f"{API}/rentals/export", params={"month": rent_month, "format": "pdf"})
        assert p.status_code == 200 and p.content[:4] == b"%PDF"

    def test_unknown_format_silently_returns_pdf(self, client, prop, month):
        """format=bogus is not rejected — it silently falls back to PDF (should be 400/422)."""
        r = client.get(f"{API}/mis/export",
                       params={"property_id": prop["id"], "month": month, "format": "bogus"})
        assert r.status_code in (400, 422), \
            f"unknown format accepted -> {r.status_code} {r.headers.get('content-type')}"

    def test_xlsx_requires_auth(self, prop, month):
        r = requests.get(f"{API}/mis/export",
                         params={"property_id": prop["id"], "month": month, "format": "xlsx"})
        assert r.status_code == 401

    def test_invalid_property_id(self, client, month):
        r = client.get(f"{API}/mis/export",
                       params={"property_id": "deadbeef", "month": month, "format": "xlsx"})
        assert r.status_code == 400
