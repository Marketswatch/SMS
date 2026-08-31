"""Iteration 13 — booking/delivery dates, opening dues + payer, renames,
floor->flat ordering, and the month-end report pack (/api/reports/pack).

Run serially:  python -m pytest tests/test_iter13_booking_dues_pack.py -n 0
"""
import io
import os
import zipfile

import fitz
import pytest
import requests
from dotenv import dotenv_values
from openpyxl import load_workbook

fe = dotenv_values("/app/frontend/.env")
BASE = (os.environ.get("REACT_APP_BACKEND_URL") or fe.get("REACT_APP_BACKEND_URL")).rstrip("/")
API = f"{BASE}/api"
MONTH = "2026-08"
NEXT = "2026-09"
FLOOR_ORDER = ["ground", "first", "second", "third", "fourth", "fifth"]


def frank(f):
    s = str(f or "").strip().lower()
    return FLOOR_ORDER.index(s) if s in FLOOR_ORDER else (99 if not s else 90)


def fkey(row):
    num = str(row.get("flat_number") or row.get("number") or "")
    d = "".join(c for c in num if c.isdigit())
    return (frank(row.get("floor")), int(d) if d else 0, num)


def pdf_text(data):
    d = fitz.open(stream=data, filetype="pdf")
    return "\n".join(p.get_text() for p in d), d


# ------------------------------------------------------------------ fixtures
@pytest.fixture(scope="module")
def client():
    s = requests.Session()
    r = s.post(f"{API}/auth/login", json={"email": "admin@societyhub.com", "password": "admin123"})
    assert r.status_code == 200, r.text
    tok = r.json().get("access_token") or r.json().get("token")
    if tok:
        s.headers.update({"Authorization": f"Bearer {tok}"})
    return s


@pytest.fixture(scope="module")
def prop(client):
    r = client.get(f"{API}/properties")
    assert r.status_code == 200, r.text
    props = r.json()
    named = [p for p in props if p["name"] == "Sunrise Residency"]
    assert named or props, "no maintenance property"
    return (named or props)[0]


@pytest.fixture(scope="module")
def stmt(client, prop):
    r = client.get(f"{API}/statement", params={"property_id": prop["id"], "month": MONTH})
    assert r.status_code == 200, r.text[:300]
    return r.json()


# ============================================================ tanker dates
class TestTankerDates:
    def test_post_delivery_before_booking_rejected(self, client, prop):
        body = {"property_id": prop["id"], "month": MONTH, "date": "2026-08-10",
                "booking_date": "2026-08-15", "qty_sump": 1000, "amount": 900,
                "supplier": "TEST_bad"}
        r = client.post(f"{API}/tankers", json=body)
        assert r.status_code == 400, f"expected 400, got {r.status_code}: {r.text[:200]}"
        assert "on or after the booking date" in r.json().get("detail", "")

    def test_post_and_put_booking_delivery_ok_and_put_validation(self, client, prop):
        body = {"property_id": prop["id"], "month": MONTH, "date": "2026-08-20",
                "booking_date": "2026-08-18", "qty_sump": 2000, "qty_syntex": 0,
                "amount": 1800, "supplier": "TEST_ok"}
        r = client.post(f"{API}/tankers", json=body)
        assert r.status_code == 200, r.text[:300]
        doc = r.json()
        tid = doc["id"]
        try:
            assert doc["booking_date"] == "2026-08-18"
            assert doc["date"] == "2026-08-20"
            assert doc["month"] == MONTH
            # GET must show it
            lst = client.get(f"{API}/tankers", params={"property_id": prop["id"], "month": MONTH}).json()
            got = [t for t in lst if t["id"] == tid]
            assert got and got[0]["booking_date"] == "2026-08-18"

            bad = dict(body, date="2026-08-01")
            r2 = client.put(f"{API}/tankers/{tid}", json=bad)
            assert r2.status_code == 400, f"PUT expected 400, got {r2.status_code}"
            assert "on or after the booking date" in r2.json().get("detail", "")
        finally:
            client.delete(f"{API}/tankers/{tid}")

    def test_delivery_month_drives_filing_and_reserve(self, client, prop):
        """Booked in Aug, delivered in Sep -> purchase belongs to September."""
        before_aug = client.get(f"{API}/statement",
                                params={"property_id": prop["id"], "month": MONTH}).json()["totals"]
        before_sep = client.get(f"{API}/statement",
                                params={"property_id": prop["id"], "month": NEXT}).json()["totals"]
        body = {"property_id": prop["id"], "month": MONTH, "date": "2026-09-03",
                "booking_date": "2026-08-28", "qty_sump": 5000, "qty_syntex": 0,
                "amount": 4500, "supplier": "TEST_crossmonth"}
        r = client.post(f"{API}/tankers", json=body)
        assert r.status_code == 200, r.text[:300]
        tid = r.json()["id"]
        try:
            assert r.json()["month"] == NEXT, "delivery month must override the posted month"
            aug = client.get(f"{API}/tankers", params={"property_id": prop["id"], "month": MONTH}).json()
            sep = client.get(f"{API}/tankers", params={"property_id": prop["id"], "month": NEXT}).json()
            assert tid not in [t["id"] for t in aug], "purchase still filed under booking month"
            assert tid in [t["id"] for t in sep]

            after_aug = client.get(f"{API}/statement",
                                   params={"property_id": prop["id"], "month": MONTH}).json()["totals"]
            after_sep = client.get(f"{API}/statement",
                                   params={"property_id": prop["id"], "month": NEXT}).json()["totals"]
            assert after_aug["total_litres"] == before_aug["total_litres"], "August litres changed"
            assert round(after_sep["total_litres"] - before_sep["total_litres"], 2) == 5000.0
            assert round(after_sep["reserve_litres"] - before_sep["reserve_litres"], 2) == 5000.0
        finally:
            client.delete(f"{API}/tankers/{tid}")


# ============================================================ opening dues
class TestOpeningDues:
    def test_flat_opening_dues_flow_to_statement(self, client, prop):
        r = client.post(f"{API}/flats", json={"property_id": prop["id"], "number": "TEST901",
                                              "floor": "Second", "owner_name": "TEST_Dues Owner",
                                              "tenant_name": "TEST_Tenant",
                                              "opening_dues": 5000, "opening_dues_payer": "tenant"})
        assert r.status_code == 200, r.text[:300]
        flat = r.json()
        assert flat["opening_dues"] == 5000
        assert flat["opening_dues_payer"] == "tenant"
        fid = flat["id"]
        try:
            got = [f for f in client.get(f"{API}/flats", params={"property_id": prop["id"]}).json()
                   if f["id"] == fid]
            assert got and got[0]["opening_dues"] == 5000 and got[0]["opening_dues_payer"] == "tenant"

            s = client.get(f"{API}/statement", params={"property_id": prop["id"], "month": MONTH}).json()
            row = next((x for x in s["rows"] if x["flat_id"] == fid), None)
            assert row is not None, "new flat missing from statement"
            assert row["carry_in"] == 5000.0, f"carry_in={row['carry_in']}"
            assert row["carry_in_payer"] == "tenant", f"carry_in_payer={row['carry_in_payer']}"
            assert row["opening_dues"] == 5000.0
            # formula: total + carry - advance - paid = balance
            expect = round(row["base_cost"] + row["carry_in"] - row["contributions"]
                           - row["received"] + row["payouts"], 2)
            assert abs(row["net"] - expect) < 0.02, f"net {row['net']} != formula {expect}"
        finally:
            client.delete(f"{API}/flats/{fid}")

    def test_reset_does_not_double_count_opening_dues(self, client):
        """Scratch property: close month -> next month's carry_in = closed net, not net + dues."""
        p = client.post(f"{API}/properties", json={"name": "TEST_ZZ Dues Property"})
        assert p.status_code == 200, p.text[:300]
        pid = p.json()["id"]
        try:
            f = client.post(f"{API}/flats", json={"property_id": pid, "number": "101", "floor": "Ground",
                                                  "owner_name": "TEST_A", "opening_dues": 5000,
                                                  "opening_dues_payer": "tenant"})
            assert f.status_code == 200, f.text[:300]
            fid = f.json()["id"]
            m = "2026-08"
            s1 = client.get(f"{API}/statement", params={"property_id": pid, "month": m}).json()
            row1 = s1["rows"][0]
            assert row1["carry_in"] == 5000.0
            net1 = row1["net"]

            r = client.post(f"{API}/periods/reset", params={"property_id": pid, "month": m})
            assert r.status_code == 200, r.text[:300]
            assert r.json()["new_month"] == "2026-09"
            assert abs(float(r.json()["carry_in"][fid]) - net1) < 0.01

            s2 = client.get(f"{API}/statement", params={"property_id": pid, "month": "2026-09"}).json()
            row2 = s2["rows"][0]
            assert abs(row2["carry_in"] - net1) < 0.01, \
                f"next-month carry_in {row2['carry_in']} should equal closed net {net1}"
            assert row2["carry_in"] != net1 + 5000, "opening dues double counted"
            assert row2["carry_in_payer"] == "", "carry_in_payer should be blank once a month is closed"
        finally:
            client.delete(f"{API}/properties/{pid}")


# ============================================================ ordering
class TestOrdering:
    def test_flats_sorted_floor_then_flat(self, client, prop):
        flats = client.get(f"{API}/flats", params={"property_id": prop["id"]}).json()
        keys = [fkey(f) for f in flats]
        assert keys == sorted(keys), [(f.get("floor"), f.get("number")) for f in flats]

    def test_statement_rows_and_meters_sorted(self, stmt):
        keys = [fkey(r) for r in stmt["rows"]]
        assert keys == sorted(keys), [(r.get("floor"), r.get("flat_number")) for r in stmt["rows"]]
        mkeys = [(fkey(m), str(m.get("label") or "")) for m in stmt["meters"]]
        assert mkeys == sorted(mkeys), [(m.get("floor"), m.get("flat_number"), m.get("label"))
                                        for m in stmt["meters"]]

    def test_statement_row_has_paid_by(self, stmt):
        for r in stmt["rows"]:
            assert "last_paid_by" in r
            if r.get("last_paid_on"):
                assert r["last_paid_by"] in ("owner", "tenant"), r["last_paid_by"]


# ============================================================ report pack
class TestReportPack:
    def test_combined_pdf(self, client, prop, stmt):
        r = client.get(f"{API}/reports/pack", params={"property_id": prop["id"], "month": MONTH,
                                                      "report": "all", "format": "pdf"})
        assert r.status_code == 200, r.text[:300]
        assert r.headers["content-type"].startswith("application/pdf")
        text, doc = pdf_text(r.content)
        assert doc.page_count >= 2
        assert "Month-end owner report pack" in text, "cover page missing"
        for title in ["as per meter readings", "Total water purchases for the month",
                      "Recurring entries", "Water reconciliation"]:
            assert title in text, f"section missing: {title}"
        assert "Advance payment" in text and "paid by" in text
        assert "Booking" in text and "Delivery" in text
        assert text.count("Colour key") >= 4, f"colour keys={text.count('Colour key')}"
        # DD-MM-YYYY dates
        import re
        assert re.search(r"\b\d{2}-\d{2}-20\d{2}\b", text), "no DD-MM-YYYY date in pack"
        # totals match statement
        t = stmt["totals"]
        assert f"{t['billable_total']:,.2f}" in text, "billable total missing from pack"
        assert f"{t['total_litres']:,.2f}" in text, "total litres missing from pack"

    @pytest.mark.parametrize("rep,marker,absent", [
        ("meters", "as per meter readings", "Water reconciliation"),
        ("purchases", "Total water purchases", "Recurring entries"),
        ("recurring", "Recurring entries", "Total water purchases"),
        ("reconciliation", "Water reconciliation", "as per meter readings"),
    ])
    def test_single_report_pdf(self, client, prop, rep, marker, absent):
        r = client.get(f"{API}/reports/pack", params={"property_id": prop["id"], "month": MONTH,
                                                      "report": rep, "format": "pdf"})
        assert r.status_code == 200, r.text[:300]
        text, _ = pdf_text(r.content)
        assert marker in text
        assert absent not in text, f"{rep} PDF leaked another report"
        assert "Month-end owner report pack" not in text, "single report should have no cover"

    def test_png(self, client, prop):
        r = client.get(f"{API}/reports/pack", params={"property_id": prop["id"], "month": MONTH,
                                                      "report": "all", "format": "png"})
        assert r.status_code == 200, r.text[:300]
        assert r.headers["content-type"] == "image/png"
        from PIL import Image
        im = Image.open(io.BytesIO(r.content))
        assert im.width > 500 and im.height > im.width, f"{im.size} not a tall stacked sheet"

    def test_zip(self, client, prop):
        r = client.get(f"{API}/reports/pack", params={"property_id": prop["id"], "month": MONTH,
                                                      "report": "all", "format": "zip"})
        assert r.status_code == 200, r.text[:300]
        assert r.headers["content-type"] == "application/zip"
        z = zipfile.ZipFile(io.BytesIO(r.content))
        names = z.namelist()
        assert len(names) == 9, names
        for rep in ("meters", "purchases", "recurring", "reconciliation"):
            assert any(n.startswith(rep) and n.endswith(".pdf") for n in names), rep
            assert any(n.startswith(rep) and n.endswith(".png") for n in names), rep
        assert any(n.startswith("month-end-pack") and n.endswith(".pdf") for n in names)
        for n in names:
            assert len(z.read(n)) > 1000, f"{n} too small"

    @pytest.mark.parametrize("params", [
        {"report": "all", "format": "docx"},
        {"report": "nonsense", "format": "pdf"},
    ])
    def test_bad_params_400(self, client, prop, params):
        r = client.get(f"{API}/reports/pack", params={"property_id": prop["id"], "month": MONTH, **params})
        assert r.status_code == 400, f"expected 400, got {r.status_code}: {r.text[:200]}"

    def test_pdf_colours_navy_header_and_unique_rows(self, client, prop, stmt):
        """Rasterise the reconciliation page: navy header band, distinct per-owner row fills."""
        r = client.get(f"{API}/reports/pack", params={"property_id": prop["id"], "month": MONTH,
                                                      "report": "reconciliation", "format": "pdf"})
        assert r.status_code == 200
        doc = fitz.open(stream=r.content, filetype="pdf")
        drawings = doc[0].get_drawings()
        fills = [d["fill"] for d in drawings if d.get("fill")]
        hexes = {"#%02x%02x%02x" % tuple(int(round(c * 255)) for c in f) for f in fills}
        assert "#1f3864" in hexes, f"navy header fill missing; fills={sorted(hexes)[:12]}"
        # one distinct tint per owner row (plus total band + navy)
        owners = {r_["owner_name"] for r_ in stmt["rows"]}
        assert len(hexes) >= len(owners) + 1, \
            f"only {len(hexes)} fill colours for {len(owners)} owners: {sorted(hexes)}"

    def test_nothing_clipped(self, client, prop):
        """Every drawing/text block must sit inside the printable page area."""
        r = client.get(f"{API}/reports/pack", params={"property_id": prop["id"], "month": MONTH,
                                                      "report": "all", "format": "pdf"})
        doc = fitz.open(stream=r.content, filetype="pdf")
        margin = 12 * 72 / 25.4  # 12 mm doc margin
        problems = []
        for i, page in enumerate(doc):
            w, h = page.rect.width, page.rect.height
            for b in page.get_text("blocks"):
                x0, y0, x1, y1 = b[:4]
                if x0 < -1 or y0 < -1 or x1 > w + 1 or y1 > h + 1:
                    problems.append(f"page {i + 1}: text {b[4][:25]!r} outside {w:.0f}x{h:.0f}")
            widest = max((d["rect"].x1 for d in page.get_drawings()), default=0)
            if widest > w - margin + 1:
                problems.append(f"page {i + 1}: table right edge {widest:.1f} exceeds "
                                f"printable width {w - margin:.1f} (page {w:.1f})")
        assert not problems, problems


# ============================================================ MIS regression
class TestMisRegression:
    @pytest.mark.parametrize("fmt", ["csv", "pdf", "xlsx"])
    def test_export_ok(self, client, prop, fmt):
        r = client.get(f"{API}/mis/export", params={"property_id": prop["id"], "month": MONTH,
                                                    "format": fmt})
        assert r.status_code == 200, r.text[:200]
        assert len(r.content) > 500

    def test_csv_renamed_and_new_columns(self, client, prop):
        r = client.get(f"{API}/mis/export", params={"property_id": prop["id"], "month": MONTH,
                                                    "format": "csv"})
        txt = r.content.decode("utf-8", "ignore")
        assert "Advance payment paid by" in txt
        assert "Fronted by owners" not in txt and "Advance paid (fronting)" not in txt
        assert "Paid by" in txt
        # NOTE: the CSV/PDF MIS exports carry no tanker table (xlsx only), so no
        # Booking/Delivery columns are expected here.

    def test_pdf_renamed_columns(self, client, prop):
        r = client.get(f"{API}/mis/export", params={"property_id": prop["id"], "month": MONTH,
                                                    "format": "pdf"})
        text, _ = pdf_text(r.content)
        assert "Advance payment" in text
        assert "Advance paid (fronting)" not in text and "Fronted by owners" not in text
        assert "Paid by" in text or "Paid\nby" in text

    def test_xlsx_tanker_and_recon_columns(self, client, prop):
        r = client.get(f"{API}/mis/export", params={"property_id": prop["id"], "month": MONTH,
                                                    "format": "xlsx"})
        wb = load_workbook(io.BytesIO(r.content))
        flat = "\n".join(str(c.value) for ws in wb.worksheets for row in ws.iter_rows()
                         for c in row if c.value is not None)
        assert "Advance payment paid by" in flat
        assert "Advance paid (fronting)" not in flat and "Fronted by owners" not in flat
        tank = wb["Tanker Purchases"]
        heads = [str(c.value) for row in tank.iter_rows(max_row=12) for c in row if c.value]
        assert any("Booking" in h for h in heads), heads
        assert any("Delivery" in h for h in heads), heads

    def test_annual_export(self, client, prop):
        r = client.get(f"{API}/annual/export", params={"property_id": prop["id"], "year": "2026",
                                                       "format": "csv"})
        assert r.status_code == 200, r.text[:200]
        assert len(r.content) > 200
